import csv
from io import TextIOWrapper, StringIO, BytesIO


from flask import request, jsonify, send_file, current_app
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation


from flask import Flask, render_template, request, redirect, url_for, flash, Response
from flask_mysqldb import MySQL,MySQLdb

from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta

# # Para subir archivos tipo foto al servidor
import os
import uuid
from werkzeug.utils import secure_filename
from flask import send_from_directory

from flask import Flask, render_template, request, jsonify 

from email_service import send_mantenimiento_notification_html


from flask import Blueprint, render_template, redirect, url_for, request, flash
from flask_login import login_user, logout_user, login_required, current_user

from extensions import db, login_manager
from models.entities.User import User
from models.ModelUser import ModelUser
from datetime import datetime


bp = Blueprint('main', __name__)

UPLOAD_FOLDER_PDF = "/var/www/sistema_mantenimientos_tecnologia/static/pdf"
ALLOWED_PDF = {"pdf"}
UPLOAD_FOLDER = "/var/www/sistema_mantenimientos_tecnologia/static/fotos"
ALLOWED_EXTENSIONS = {".png", ".jpg", ".jpeg"}

# @bp.context_processor
# def link_onedrive_mantenimiento():
#     return dict(onedrive_link_mantenimiento=LinkOneDriveMantenimiento.ONEDRIVE_LINK_MANTENIMIENTO)

# @bp.context_processor
# def link_onedrive_calibracion():
#     return dict(onedrive_link_calibracion=LinkOneDriveCalibracion.ONEDRIVE_LINK_CALIBRACION)

@login_manager.user_loader
def load_user(id):
    return ModelUser.get_by_id(db, id)

@bp.after_request
def evita_cache(response):
    response.cache_control.no_store = True
    response.cache_control.no_cache = True
    response.cache_control.must_revalidate = True
    response.cache_control.max_age = 0
    response.expires = 0
    response.pragma = 'no-cache'
    return response

@bp.route('/')
# @login_required
def index():
    return redirect(url_for('login'))


@bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User(0, request.form['username'], request.form['password'])
        logged_user = ModelUser.login(db, user)

        if logged_user:
            if logged_user.password:
                login_user(logged_user)

                # Redirección directa por rol
                if logged_user.rol == 'tecnologia':
                    return redirect(url_for('main.indexTecnologia'))
                
                elif logged_user.rol == 'admin':
                    return redirect(url_for('main.home'))
                else:
                    flash('Rol no autorizado')
                    return redirect(url_for('login'))
            else:
                flash("Contraseña incorrecta...")
        else:
            flash("Usuario no encontrado...")
    return render_template('auth/login.html')


@bp.route('/logout')
# @login_required
def logout():
    logout_user()
    return redirect(url_for('main.login'))

@bp.route('/home')
@login_required
def home():
    cur = db.connection.cursor()

    # Obtener la fecha actual
    fecha_actual = datetime.now().date()

    # Consultar la cantidad de equipos según el vencimiento del mantenimiento
    cur.execute("""
        SELECT 
            SUM(vencimiento_mantenimiento IS NULL) AS sin_fecha,
            SUM(DATEDIFF(vencimiento_mantenimiento, %s) < 0) AS vencidas,
            SUM(DATEDIFF(vencimiento_mantenimiento, %s) BETWEEN 0 AND 30) AS proximas,
            SUM(DATEDIFF(vencimiento_mantenimiento, %s) > 30) AS mas_30_dias
        FROM tecnologia_equipos WHERE enable = 1 and de_baja = 0
    """, (fecha_actual, fecha_actual, fecha_actual))

    # Obtener los resultados
    resultados_preventivos = cur.fetchone()
    sin_fecha_mantenimiento_preventivo = resultados_preventivos[0] or 0
    vencidas_mantenimiento_preventivo = resultados_preventivos[1] or 0
    entre_0_y_30_dias_a_vencer_preventivo = resultados_preventivos[2] or 0
    mas_30_dias_a_vencer_preventivo = resultados_preventivos[3] or 0

    print(resultados_preventivos)
   

    # Consultar la cantidad de equipos según el vencimiento de la calibración
    cur.execute("""
        SELECT 
            SUM(vencimiento_calibracion IS NULL) AS sin_fecha,
            SUM(DATEDIFF(vencimiento_calibracion, %s) < 0) AS vencidas,
            SUM(DATEDIFF(vencimiento_calibracion, %s) BETWEEN 0 AND 30) AS proximas,
            SUM(DATEDIFF(vencimiento_calibracion, %s) > 30) AS mas_30_dias
        FROM tecnologia_equipos WHERE enable = 1 and de_baja = 0
    """, (fecha_actual, fecha_actual, fecha_actual))

    # Obtener los resultados del vencimiento de calibración
    resultados_correctivos = cur.fetchone()
    sin_fecha_mantenimiento_correctivo = resultados_correctivos[0] or 0
    vencidas_mantenimiento_correctivo = resultados_correctivos[1] or 0
    entre_0_y_30_dias_a_vencer_correctivo = resultados_correctivos[2] or 0
    mas_30_dias_a_vencer_correctivo = resultados_correctivos[3] or 0

    return render_template('home.html', sin_fecha_mantenimiento_preventivo=sin_fecha_mantenimiento_preventivo, 
                                        vencidas_mantenimiento_preventivo=vencidas_mantenimiento_preventivo, 
                                        entre_0_y_30_dias_a_vencer_preventivo=entre_0_y_30_dias_a_vencer_preventivo, 
                                        mas_30_dias_a_vencer_preventivo=mas_30_dias_a_vencer_preventivo,
                                        
                                        sin_fecha_mantenimiento_correctivo=sin_fecha_mantenimiento_correctivo, 
                                        vencidas_mantenimiento_correctivo=vencidas_mantenimiento_correctivo, 
                                        entre_0_y_30_dias_a_vencer_correctivo=entre_0_y_30_dias_a_vencer_correctivo, 
                                        mas_30_dias_a_vencer_correctivo=mas_30_dias_a_vencer_correctivo)


# ---------------------------INICIA MODULO DE TECNOLOGIA-----------------------------
@bp.route('/datosTecnicoTecnologia/<id>')
@login_required
def DATOS_TECNICO_TECNOLOGIA(id):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    # cur = db.connection.cursor()
    cur.execute('SELECT * FROM tecnologia_tecnico_responsable WHERE id = %s', [id])
    proveedor = cur.fetchall()
    print (proveedor)
    return render_template('datosTecnicoTecnologia.html', tecnologia_tecnico_responsable=proveedor)

# ===============================DATOS PERSONA RESPONSABLE TECNOLOGIA================================
@bp.route('/datosPersonaTecnologia/<id>')
@login_required
def DATOS_PERSONA_TECNOLOGIA(id):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    # cur = db.connection.cursor()
    cur.execute('SELECT * FROM tecnologia_persona_responsable WHERE id = %s', [id])
    persona = cur.fetchall()
    print (persona)
    return render_template('datosPersonaTecnologia.html', tecnologia_persona_responsable=persona)

# FUNCIÓN ACTUALIZAR DATOS PERSONAS
@bp.route('/update_datos_persona_tecnologia/<id>', methods = ['POST'])
def ACTUALIZAR_DATOS_PERSONA_TECNOLOGIA(id):
    if request.method =='POST':
        documento_identidad = request.form ['documento_identidad']
        # nombre_empresa = request.form ['nombre_empresa']
        nombre_contratista = request.form ['nombre_contratista']
        correo = request.form ['correo']
        # cargo_contacto = request.form ['cargo_contacto']
        area = request.form ['area']
        cur = db.connection.cursor() 
        cur.execute(""" UPDATE tecnologia_persona_responsable SET documento_identidad = %s, nombre_contratista = %s, correo = %s, area = %s WHERE id = %s """, 
                                                      (documento_identidad, nombre_contratista, correo, area, id))
        db.connection.commit()
    flash('Datos actualizados satisfactorimanete', 'success')
    return redirect(url_for('main.indexTecnologia', id = id))
    # return redirect(url_for('datosPersonaTecnologia')) 

# ESTA FUNCIÓN ME LLEVA A OTRA VISTA PARA AGREGAR LOS NUEVAS PERSONAS
@bp.route('/agregarNuevaPersonaTecnologia')
@login_required
def AGREGAR_NUEVA_PERSONA_TECNOLOGIA():
    return render_template('agregarNuevaPersonaTecnologia.html')

@bp.route('/add_datosPersonaTecnologia', methods=['POST'])
def EDITAR_DATOS_PERSONA_TECNOLOGIA():
    if request.method == 'POST':
        documento_identidad = request.form.get('documento_identidad')
        nombre_contratista = request.form.get('nombre_contratista')
        correo = request.form.get('correo')
        area = request.form.get('area')

        # ✅ Valida que todos los campos estén diligenciados
        if not documento_identidad or not nombre_contratista or not correo or not area:
            flash('Todos los campos son obligatorios', 'danger')
            return redirect(url_for('main.AGREGAR_NUEVA_PERSONA_TECNOLOGIA'))

        cur = db.connection.cursor()

        # ✅ Verifica si ya existe el documento_identidad antes de insertar
        cur.execute(
            "SELECT COUNT(*) FROM tecnologia_persona_responsable WHERE documento_identidad = %s",
            (documento_identidad,)
        )
        existe = cur.fetchone()[0]

        if existe:
            flash('El documento de identidad ya está registrado', 'warning')
            return redirect(url_for('main.AGREGAR_NUEVA_PERSONA_TECNOLOGIA'))

        # ✅ Inserta solo si no existe
        cur.execute(
            'INSERT INTO tecnologia_persona_responsable (documento_identidad, nombre_contratista, correo, area) '
            'VALUES (%s, %s, %s, %s)',
            (documento_identidad, nombre_contratista, correo, area)
        )
        db.connection.commit()

        flash('Datos agregados satisfactoriamente', 'success')
        return redirect(url_for('main.AGREGAR_NUEVA_PERSONA_TECNOLOGIA'))
# ============================================================================================

@bp.route('/indexTecnologia')
@login_required
def indexTecnologia():
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    # cur = db.connection.cursor()
    cur.execute('SELECT * FROM tecnologia_equipos where enable=1 AND de_baja=0 AND otros_equipos_tecnologia = 0')
    data = cur.fetchall()
    # cur.execute('SELECT i.*, p.enable_prestamos FROM tecnologia_equios i LEFT JOIN prestamos_equiposalud p ON i.cod_articulo = p.cod_articulo AND p.enable_prestamos = 1 WHERE i.enable=1 AND i.de_baja=0') #A raiz del enable=1 no se deben eliminar en la DB
    # data = cur.fetchall()

    cur.execute('SELECT id, nombre_tecnico FROM tecnologia_tecnico_responsable')
    proveedores = cur.fetchall()

    cur.execute('SELECT id, documento_identidad, nombre_contratista FROM tecnologia_persona_responsable')
    personas = cur.fetchall()

    cur.execute('SELECT id, tipo_equipo FROM tecnologia_tipo_equipo')
    tipoEquipos = cur.fetchall()

    cur.execute('SELECT id, estado_equipo FROM tecnologia_estados_equipos')
    estadoEquipos = cur.fetchall()

    cur.execute('SELECT id, proceso FROM tecnologia_procesos WHERE activo = 1 ORDER BY proceso')
    procesoEquipos = cur.fetchall()
    
    cur.execute('SELECT id, proceso FROM tecnologia_procesos')
    procesoEquipos_data = cur.fetchall()
    procesoEquiposModal = {p["id"]: p["proceso"] for p in procesoEquipos_data}
    
    # print(procesoEquipos)
    return render_template('indexTecnologia.html', tecnologia_equipos=data, tipoEquipos=tipoEquipos, proveedores=proveedores, personas=personas, estadoEquipos=estadoEquipos, procesoEquipos=procesoEquipos, procesoEquiposModal=procesoEquiposModal)


def allowed_image(filename):
    ext = os.path.splitext(filename.lower())[1]
    return ext in ALLOWED_EXTENSIONS


@bp.route('/add_equipos_tecnologia', methods=['GET', 'POST'])
def add_equipos_tecnologia():
    if request.method == 'POST':
        # ===== VALIDAR CAMPOS =====
        cod_articulo = request.form.get('cod_articulo')
        nombre_equipo = request.form.get('nombre_equipo')

        try:
            cod_articulo = int(cod_articulo)
        except:
            flash("Por favor ingresar solo números en el código del equipo.", "error")
            return redirect(url_for('main.indexTecnologia'))

        # ===== VERIFICAR EXISTENCIA =====
        cur = db.connection.cursor()
        cur.execute("SELECT 1 FROM tecnologia_equipos WHERE cod_articulo = %s", (cod_articulo,))
        if cur.fetchone():
            flash(f"El código {cod_articulo} ya existe.", "error")
            return redirect(url_for('main.indexTecnologia'))

        # ===== MANEJO DE IMAGEN (OPCIONAL) =====
        DEFAULT_IMAGE = "fotos/pcs-animado.jpg"
        image_path_db = DEFAULT_IMAGE

        file = request.files.get('imagen_producto')

        if file and file.filename:
            if not allowed_image(file.filename):
                flash("Formato inválido. Solo se permiten PNG, JPG, JPEG.", "error")
                return redirect(url_for('main.indexTecnologia'))

            os.makedirs(UPLOAD_FOLDER, exist_ok=True)

            ext = os.path.splitext(file.filename)[1].lower()
            unique_name = f"{uuid.uuid4().hex}{ext}"
            save_path = os.path.join(UPLOAD_FOLDER, unique_name)

            try:
                file.save(save_path)
                image_path_db = f"fotos/{unique_name}"
            except Exception as e:
                flash("Error al guardar la imagen. Se usará imagen por defecto.", "warning")
                image_path_db = DEFAULT_IMAGE

        # ===== RESTO DE CAMPOS (limpios) =====
        fecha_ingreso = request.form.get("fecha_ingreso") or None
        estado_equipo = request.form.get("estado_equipo")

        # ===== MANEJO DE PDF DE BAJA (SOLO SI ES DE BAJA) =====
        pdf_path_db = None
        
        if estado_equipo == "DE BAJA":
            pdf_file = request.files.get("pdf_debaja")

            if not pdf_file or not pdf_file.filename:
                flash("Debe adjuntar el acta de baja en PDF.", "error")
                return redirect(url_for("main.indexTecnologia"))

            if not allowed_pdf(pdf_file.filename):
                flash("El documento debe ser PDF.", "error")
                return redirect(url_for("main.indexTecnologia"))

            os.makedirs(UPLOAD_FOLDER_PDF, exist_ok=True)

            # ✅ (recomendado) conservar nombre real y evitar colisiones
            original_name = secure_filename(pdf_file.filename)
            base, ext = os.path.splitext(original_name)
            final_name = original_name
            counter = 1
            while os.path.exists(os.path.join(UPLOAD_FOLDER_PDF, final_name)):
                final_name = f"{base}_{counter}{ext}"
                counter += 1

            save_path = os.path.join(UPLOAD_FOLDER_PDF, final_name)
            pdf_file.save(save_path)

            pdf_path_db = f"pdf/{final_name}"

        enable = 0
        de_baja = 0
        otros_equipos_tecnologia = 0

        if estado_equipo in ("USO", "SIN USO"):
            enable = 1

        elif estado_equipo == "DE BAJA":
            de_baja = 1

        elif estado_equipo == "OTROS EQUIPOS":
            otros_equipos_tecnologia = 1

        fecha_de_baja = date.today() if estado_equipo == "DE BAJA" else None

        # ===== INSERT EN tecnología_equipos =====
        cur.execute("""
            INSERT INTO tecnologia_equipos (
                cod_articulo, nombre_equipo, fecha_ingreso, tipo_equipo, 
                estado_equipo, id_proceso, ram, disco, 
                color, checkbox_mantenimiento, checkbox_calibracion, imagen, 
                software_instalado, marca_equipo_tecnologia, modelo_equipo_tecnologia, 
                serial_equipo_tecnologia, id_persona_responsable, enable, otros_equipos_tecnologia, ubicacion, de_baja
            )
            VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            cod_articulo,
            nombre_equipo,
            fecha_ingreso,
            request.form.get("tipo_equipo"),
            estado_equipo,
            request.form.get("id_proceso"),
            request.form.get("ram"),
            request.form.get("disco"),
            # request.form.get("proveedor_responsable"),
            "verde",   # color fijo que estabas usando
            "Inactivo",
            "Inactivo",
            image_path_db,
            request.form.get("software_instalado"),
            request.form.get("marca_equipo_tecnologia"),
            request.form.get("modelo_equipo_tecnologia"),
            request.form.get("serial_equipo_tecnologia"),
            request.form.get("id_persona_responsable"),
            enable,
            otros_equipos_tecnologia,
            request.form.get("ubicacion"),
            de_baja
        ))

        # ===== INSERT EN DE BAJA =====
        if estado_equipo == "DE BAJA":
            cur.execute("""
                INSERT INTO tecnologia_equipos_debaja (
                    cod_articulo, nombre_equipo, fecha_ingreso, tipo_equipo,
                    estado_equipo, id_proceso, ram, disco, 
                    color, checkbox_mantenimiento, checkbox_calibracion, imagen, 
                    software_instalado, marca_equipo_tecnologia, modelo_equipo_tecnologia, 
                    serial_equipo_tecnologia, fecha_de_baja, id_persona_responsable, pdf_debaja
                )
                VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                cod_articulo,
                nombre_equipo,
                fecha_ingreso,
                request.form.get("tipo_equipo"),
                estado_equipo,
                request.form.get("id_proceso"),
                request.form.get("ram"),
                request.form.get("disco"),
                # request.form.get("proveedor_responsable"),
                "verde",
                "Inactivo",
                "Inactivo",
                image_path_db,
                request.form.get("software_instalado"),
                request.form.get("marca_equipo_tecnologia"),
                request.form.get("modelo_equipo_tecnologia"),
                request.form.get("serial_equipo_tecnologia"),
                fecha_de_baja,
                request.form.get("id_persona_responsable"),
                pdf_path_db
            ))

        db.connection.commit()

        flash("Equipo agregado correctamente", "success")

        if estado_equipo == "OTROS EQUIPOS":
            return redirect(url_for('main.index_otros_equipos_tecnologia'))

        return redirect(url_for('main.indexTecnologia'))

    return render_template('indexTecnologia.html')
# ---------------------------INICIA INSERT MASIVO DE EQUIPOS EXCEL DE TECNOLOGIA-----------------------------
# --------------------------- INSERT MASIVO EXCEL TECNOLOGIA ---------------------------
@bp.route('/insert_excelTecnologia', methods=['POST'])
def insert_excel_tecnologia():

    DEFAULT_IMAGE = "fotos/pcs-animado.jpg"

    file = request.files.get('file')

    if not file or not file.filename.endswith('.xlsx'):
        flash("Debe subir un archivo Excel (.xlsx)", "error")
        return redirect(url_for('main.indexTecnologia'))

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    cur = db.connection.cursor()

    # ===================================================
    # CARGAR MAPS UNA SOLA VEZ (BUENA PRÁCTICA)
    # ===================================================

    # Personas
    cur.execute("SELECT id, documento_identidad FROM tecnologia_persona_responsable")
    persona_map = {str(p[1]).strip().lower(): p[0] for p in cur.fetchall()}

    # Procesos
    cur.execute("SELECT id, proceso FROM tecnologia_procesos")
    proceso_map = {str(p[1]).strip().lower(): p[0] for p in cur.fetchall()}

    codigos_duplicados = []
    insertados = 0

    # ===================================================
    # ITERAR FILAS
    # ===================================================

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        try:

            # ===============================
            # VALIDACIONES BÁSICAS
            # ===============================

            if not row[0]:
                continue

            cod_articulo = int(row[0])
            nombre_equipo = row[1]

            # Duplicados
            cur.execute("SELECT 1 FROM tecnologia_equipos WHERE cod_articulo=%s", (cod_articulo,))
            if cur.fetchone():
                codigos_duplicados.append(str(cod_articulo))
                continue

            fecha_ingreso = row[2]
            tipo_equipo = row[3]
            estado_equipo = str(row[4]).strip().upper()

            # ===============================
            # PERSONA
            # ===============================

            persona_texto = str(row[5]).strip().lower()
            persona_id = persona_map.get(persona_texto)

            if not persona_id:
                flash(f"Fila {i}: persona '{row[5]}' no existe", "error")
                continue

            # ===============================
            # PROCESO (CONVERSIÓN TEXTO → ID)
            # ===============================

            proceso_texto = str(row[6]).strip().lower()
            id_proceso = proceso_map.get(proceso_texto)

            if not id_proceso:
                flash(f"Fila {i}: proceso '{row[6]}' no existe", "error")
                continue

            # ===============================
            # RESTO CAMPOS
            # ===============================
            ubicacion = row[7]
            ram = row[8]
            disco = row[9]
            marca = row[10]
            modelo = row[11]
            serial = row[12]
            software = row[13]

            imagen = secure_filename(row[14]) if row[14] else None
            ruta_imagen = f"fotos/{imagen}" if imagen else DEFAULT_IMAGE

            checkbox_mantenimiento = "Inactivo"
            checkbox_calibracion = "Inactivo"
            color = "verde"

            fecha_de_baja = date.today() if estado_equipo == "DE BAJA" else None

            # ===============================
            # BANDERAS ESTADO
            # ===============================

            if estado_equipo in ("USO", "SIN USO"):
                enable = 1
                de_baja = 0
                otros = 0

            elif estado_equipo == "DE BAJA":
                enable = 0
                de_baja = 1
                otros = 0

            elif estado_equipo == "OTROS EQUIPOS":
                enable = 0
                de_baja = 0
                otros = 1

            else:
                flash(f"Fila {i}: estado '{estado_equipo}' no válido", "error")
                continue

            # ===================================================
            # INSERT PRINCIPAL (AHORA CON id_proceso)
            # ===================================================

            cur.execute("""
                INSERT INTO tecnologia_equipos (
                    cod_articulo, nombre_equipo,
                    fecha_ingreso, tipo_equipo,
                    estado_equipo, id_proceso,
                    ubicacion, ram, disco,
                    color, checkbox_mantenimiento, checkbox_calibracion,
                    imagen, software_instalado,
                    marca_equipo_tecnologia,
                    modelo_equipo_tecnologia,
                    serial_equipo_tecnologia,
                    id_persona_responsable,
                    enable, de_baja, otros_equipos_tecnologia
                )
                VALUES (%s,%s,%s,%s,%s,%s,
                        %s,%s,%s,%s,%s,%s,%s,
                        %s,%s,%s,%s,
                        %s,%s,%s,%s)
            """, (
                cod_articulo,
                nombre_equipo,
                fecha_ingreso,
                tipo_equipo,
                estado_equipo,
                id_proceso,
                ubicacion,
                ram,
                disco,
                color,
                checkbox_mantenimiento,
                checkbox_calibracion,
                ruta_imagen,
                software,
                marca,
                modelo,
                serial,
                persona_id,
                enable,
                de_baja,
                otros
            ))

            # ===================================================
            # INSERT EN DE BAJA
            # ===================================================

            if estado_equipo == "DE BAJA":

                cur.execute("""
                    INSERT INTO tecnologia_equipos_debaja (
                        cod_articulo, nombre_equipo,
                        fecha_ingreso, tipo_equipo,
                        estado_equipo, id_proceso,
                        ram, disco,
                        color, checkbox_mantenimiento, checkbox_calibracion,
                        imagen, software_instalado,
                        marca_equipo_tecnologia,
                        modelo_equipo_tecnologia,
                        serial_equipo_tecnologia,
                        fecha_de_baja,
                        id_persona_responsable
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,
                            %s,%s,%s,%s,%s,%s,%s,
                            %s,%s,%s,%s,%s)
                """, (
                    cod_articulo,
                    nombre_equipo,
                    fecha_ingreso,
                    tipo_equipo,
                    estado_equipo,
                    id_proceso,
                    ram,
                    disco,
                    color,
                    checkbox_mantenimiento,
                    checkbox_calibracion,
                    ruta_imagen,
                    software,
                    marca,
                    modelo,
                    serial,
                    fecha_de_baja,
                    persona_id
                ))

            insertados += 1

        except Exception as e:
            flash(f"Fila {i}: error -> {str(e)}", "error")

    db.connection.commit()

    # ===================================================
    # MENSAJES
    # ===================================================

    if codigos_duplicados:
        flash(f"Códigos duplicados no insertados: {', '.join(codigos_duplicados)}", "warning")

    if insertados:
        flash(f"{insertados} equipos importados correctamente.", "success")
    else:
        flash("No se insertó ningún equipo.", "error")

    return redirect(url_for('main.indexTecnologia'))
# ---------------------------FINALIZA INSERT MASIVO CSV DE SALUD-----------------------------
    
# ================================CHECKBOX PROGRAMACIÓN MANTENIMIENTO TECNOLOGIA===============================
@bp.route('/checkbox_programacionMantenimientoTecnologia', methods=['POST'])
def checkbox_programacion_mantenimiento_tecnologia():
    try:
        seleccionados = request.form.getlist('seleccionados[]')
        proveedor_id = request.form.get('proveedor_id')
        persona_id = request.form.get('persona_id')
        proceso_id = request.form.get('proceso_id')

        if not seleccionados or not proveedor_id:
            return jsonify({'success': False, 'message': 'Faltan productos seleccionados o proveedor.'})

        hoy = datetime.now()
        cur = db.connection.cursor()

        productos_guardados = []

        for cod in seleccionados:
            nombre_equipo = request.form.get(f'nombre_equipo_{cod}')
            # proceso_id = request.form.get(f'proceso_{cod}')
            periodicidad_m = request.form.get(f'periodicidad_mantenimiento_{cod}')
            periodicidad_c = request.form.get(f'periodicidad_calibracion_{cod}')

            mantenimiento_activado = request.form.get(f'mantenimiento_{cod}') == 'on'
            calibracion_activada = request.form.get(f'calibracion_{cod}') == 'on'

            # Obtener fechas reales desde base de datos
            cur.execute("SELECT fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion FROM tecnologia_equipos WHERE cod_articulo = %s", (cod,))
            resultado = cur.fetchone()
            if not resultado:
                continue  # Saltar si no se encuentra

            fecha_m, vencimiento_m, fecha_c, vencimiento_c = resultado

            if mantenimiento_activado:
                # Validación de vencimiento
                if vencimiento_m and (vencimiento_m - hoy).days < 30:
                    continue
                # Activar checkbox y guardar en historial preventivo
                cur.execute(
                    "UPDATE tecnologia_equipos SET checkbox_mantenimiento = 'Activo' WHERE cod_articulo = %s",
                    (cod,),
                )
                cur.execute(
                    """INSERT INTO tecnologia_historial_preventivo 
                            (cod_articulo, nombre_equipo, id_proceso, fecha_mantenimiento, vencimiento_mantenimiento, periodicidad, id_proveedor_responsable, id_persona_responsable) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        cod,
                        nombre_equipo,
                        proceso_id,
                        fecha_m,
                        vencimiento_m,
                        periodicidad_m,
                        proveedor_id,
                        persona_id,

                    ),
                )

            if calibracion_activada:
                if vencimiento_c and (vencimiento_c - hoy).days < 30:
                    continue
                # Activar checkbox y guardar en historial correctivo
                cur.execute(
                    "UPDATE tecnologia_equipos SET checkbox_calibracion = 'Activo' WHERE cod_articulo = %s",
                    (cod,),
                )
                cur.execute(
                    """INSERT INTO tecnologia_historial_correctivo 
                            (cod_articulo, nombre_equipo, id_proceso, fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion, id_proveedor_responsable, id_persona_responsable) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        cod,
                        nombre_equipo,
                        proceso_id,
                        fecha_c,
                        vencimiento_c,
                        periodicidad_c,
                        proveedor_id,
                        persona_id,
                    ),
                )

            productos_guardados.append(cod)

        db.connection.commit()
        return jsonify({
            'success': True,
            'message': f'Se procesaron {len(productos_guardados)} productos correctamente.',
            'productos': productos_guardados
        })

    except Exception as e:
        db.connection.rollback()
        return jsonify({'success': False, 'message': f'Error en el servidor: {str(e)}'})

# OBTIENE LOS DATOS DEL RESPONSABLE DEL EQUIPO Y LLENA LOS CAMPOS DE PERSONA RESPONSABLE Y PROCESO DEL MODAL
@bp.route('/get_datos_persona/<id>', methods=['GET'])
def get_datos_persona(id):
    try:
        cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

        cur.execute("""
            SELECT 
                p.id AS persona_id,
                p.nombre_contratista,
                u.id AS proceso_id,
                u.proceso
            FROM tecnologia_equipos e
            LEFT JOIN tecnologia_persona_responsable p 
                ON e.id_persona_responsable = p.id
            LEFT JOIN tecnologia_procesos u 
                ON e.id_proceso = u.id
            WHERE e.cod_articulo = %s
            LIMIT 1
        """, (id,))

        data = cur.fetchone()
        cur.close()

        if data:
            return jsonify({'success': True, **data})
        else:
            return jsonify({'success': False, 'message': 'No se encontraron datos.'})

    except Exception as e:
        print(f"⚠️ Error en get_datos_persona: {e}")
        return jsonify({'success': False, 'message': 'Error al obtener datos.'})
    
def allowed_pdf(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_PDF
    
# ACTUALIZA EL ESTADO DEL EQUIPO DESDE EL DESPLEGABLE QUE SE ENCUENTRA EN LA MISMA TABLA INDEXSALUD
@bp.route('/update_estadoEquipoTecnologia', methods=['POST'])
@login_required
def update_estado_equipo_tecnologia():
    if request.method == 'POST':
        pdf_path_db = None
        # OBTENER FULLNAME DEL USUARIO LOGUEADO
        cur = db.connection.cursor()
        cur.execute("SELECT fullname, username FROM user WHERE id = %s", (current_user.id,))
        result = cur.fetchone()
        usuario_logueado_nombre = result[0] if result else None
        usuario_logueado_email  = result[1] if result else None

        # producto_id = request.form['producto_id']
        nuevo_estado = request.form.get("nuevo_estado_equipo")
        cod_articulo = request.form ['cod_articulo']
        nombre_equipo = request.form.get ("nombre_equipo")

        # Obtener hora actual del equipo
        hora_actual = datetime.now()

        # PARA EL CHECKBOX Y SEMAFORO DE MANTENIMIENTO
        fecha_mantenimiento = request.form.get('fecha_mantenimiento', '') or None
        vencimiento_mantenimiento = request.form.get('vencimiento_mantenimiento', '') or None
        # Convertir a objetos date solo si existen
        if fecha_mantenimiento:
            try:
                fecha_mantenimiento = datetime.strptime(fecha_mantenimiento, '%Y-%m-%d').date()
            except ValueError:
                fecha_mantenimiento = None

        if vencimiento_mantenimiento:
            try:
                vencimiento_mantenimiento = datetime.strptime(vencimiento_mantenimiento, '%Y-%m-%d').date()
            except ValueError:
                vencimiento_mantenimiento = None

        color = 'verde'
        
        # PARA EL CHECKBOX DE CALIBRACIÓN
        fecha_calibracion = request.form.get('fecha_calibracion', '') or None
        vencimiento_calibracion = request.form.get('vencimiento_calibracion', '') or None
       
        # Convertir a objetos date solo si existen
        if fecha_calibracion:
            try:
                fecha_calibracion = datetime.strptime(fecha_calibracion, '%Y-%m-%d').date()
            except ValueError:
                fecha_calibracion = None

        if vencimiento_calibracion:
            try:
                vencimiento_calibracion = datetime.strptime(vencimiento_calibracion, '%Y-%m-%d').date()
            except ValueError:
                vencimiento_calibracion = None

        fecha_ingreso = request.form.get ("fecha_ingreso")

        periodicidad_raw = request.form.get("periodicidad")
        if periodicidad_raw in (None, "", "None"):
            periodicidad = None
        else:
            periodicidad = int(periodicidad_raw)

        tipo_equipo = request.form.get ("tipo_equipo")
        # estado_equipo= request.form ['estado_equipo']

        id_proceso_raw = request.form.get ("id_proceso")
        if id_proceso_raw in (None, "", "None"):
            id_proceso = None
        else:
            id_proceso = int(id_proceso_raw)


        ram = request.form.get ("ram")
        disco = request.form.get ("disco")

        proveedor_responsable_raw = request.form.get ("proveedor_responsable")
        if proveedor_responsable_raw in (None, "", "None"):
            proveedor_responsable = None
        else:
            proveedor_responsable = int(proveedor_responsable_raw)

        software_instalado = request.form.get ("software_instalado")
        # cuidados_basicos = request.form ['cuidados_basicos']
        periodicidad_calibracion_raw = request.form.get ("periodicidad_calibracion")
        if periodicidad_calibracion_raw in (None, "", "None"):
            periodicidad_calibracion = None
        else:
            periodicidad_calibracion = int(periodicidad_calibracion_raw)

        marca_equipo_tecnologia = request.form.get ("marca_equipo_tecnologia")
        modelo_equipo_tecnologia = request.form.get ("modelo_equipo_tecnologia")
        serial_equipo_tecnologia = request.form.get ("serial_equipo_tecnologia")
        id_persona_responsable = request.form.get ("id_persona_responsable")
        cur = db.connection.cursor()

        # Obtener la ruta de la imagen desde la tabla tecnologia_equipos
        cur.execute("""
            SELECT imagen, nombre_equipo, fecha_ingreso
            FROM tecnologia_equipos
            WHERE cod_articulo = %s
        """, (cod_articulo,))
        equipo = cur.fetchone()

        if not equipo:
            flash("Equipo no encontrado.", "error")
            return redirect(url_for("main.indexTecnologia"))

        filepath_to_db_img, nombre_equipo, fecha_ingreso = equipo

        if not fecha_ingreso:
            flash("El equipo no tiene FECHA DE COMPRA registrada.", "error")
            return redirect(url_for("main.indexTecnologia"))
        
        filepath_to_db_img = equipo[0] if equipo else None

        if nuevo_estado == 'DE BAJA':

            pdf_file = request.files.get("pdf_debaja")

            if not pdf_file or not pdf_file.filename:
                flash("Debe adjuntar el acta de baja en PDF.", "error")
                return redirect(url_for("main.indexTecnologia"))

            if not allowed_pdf(pdf_file.filename):
                flash("El documento debe ser PDF.", "error")
                return redirect(url_for("main.indexTecnologia"))

            os.makedirs(UPLOAD_FOLDER_PDF, exist_ok=True)

            ext = os.path.splitext(pdf_file.filename)[1].lower()
            unique_name = f"{uuid.uuid4().hex}{ext}"

            save_path = os.path.join(UPLOAD_FOLDER_PDF, unique_name)
            pdf_file.save(save_path)

            pdf_path_db = f"pdf/{unique_name}"

            # Actualizar el estado y marcar como dado de baja en tecnologia_equipos
            cur.execute("""UPDATE tecnologia_equipos SET estado_equipo = %s, enable = 0, de_baja = 1, otros_equipos_tecnologia = 0 WHERE cod_articulo = %s""", (nuevo_estado, cod_articulo))

            # Verificar si el equipo ya está en tecnologia_equipos_debaja
            cur.execute('SELECT 1 FROM tecnologia_equipos_debaja WHERE cod_articulo = %s', (cod_articulo,))
            equipo_existente = cur.fetchone()

            # Insertar el equipo en tecnologia_equipos_debaja si no existe
            if not equipo_existente:
                cur.execute("""INSERT INTO tecnologia_equipos_debaja (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, fecha_ingreso,
                                                                periodicidad, tipo_equipo, estado_equipo, id_proceso, ram, disco, proveedor_responsable, color, imagen, software_instalado,
                                                                periodicidad_calibracion, marca_equipo_tecnologia, modelo_equipo_tecnologia, serial_equipo_tecnologia, id_persona_responsable, pdf_debaja, fecha_de_baja) 
                                                                VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        cod_articulo,
                        nombre_equipo,
                        fecha_mantenimiento,
                        vencimiento_mantenimiento,
                        fecha_calibracion,
                        vencimiento_calibracion,
                        fecha_ingreso,
                        periodicidad,
                        tipo_equipo,
                        nuevo_estado,  # Estado del equipo
                        id_proceso,
                        ram,
                        disco,
                        proveedor_responsable,
                        color,
                        filepath_to_db_img,
                        software_instalado,
                        # cuidados_basicos,
                        periodicidad_calibracion,
                        marca_equipo_tecnologia,
                        modelo_equipo_tecnologia,
                        serial_equipo_tecnologia,
                        id_persona_responsable,
                        pdf_path_db,
                        hora_actual
                    ),
                )
        
        # --- ESTADO: OTROS EQUIPOS ---
        elif nuevo_estado == 'OTROS EQUIPOS':
            cur.execute("""
                UPDATE tecnologia_equipos 
                SET estado_equipo = %s, otros_equipos_tecnologia = 1, enable = 0, de_baja = 0 
                WHERE cod_articulo = %s
            """, (nuevo_estado, cod_articulo))

        else:

        # --- 2. Actualizar el estado del equipo ---
            cur.execute("""
                UPDATE tecnologia_equipos
                SET estado_equipo = %s, otros_equipos_tecnologia = 0 
                WHERE cod_articulo = %s
            """, (nuevo_estado, cod_articulo))

        db.connection.commit()
        cur.close()
        flash('Estado del equipo actualizado correctamente', 'success')
        return redirect(url_for('main.indexTecnologia'))
# =========================================================================================================    

@bp.route('/guardar_historialTecnologia', methods=['POST'])
@login_required
def guardar_historial_tecnologia():
    data = request.get_json()
    proveedor_id = data.get('proveedorId')
    persona_id = data.get('personaId')
    proceso_id = data.get('procesoId')
    observaciones_id = data.get('observacionesId')
    nueva_fecha_str = data.get('nuevaFecha')
    correo_externo = data.get('correoExterno')
    registros = data.get('registros', [])

    try:
        if not proveedor_id or not nueva_fecha_str:
            return jsonify({'success': False, 'message': 'Falta proveedor o fecha'})

        nueva_fecha = datetime.strptime(nueva_fecha_str, '%Y-%m-%d')
        cur = db.connection.cursor()

        # --- Obtener nombres descriptivos ---
        cur.execute("SELECT nombre_tecnico FROM tecnologia_tecnico_responsable WHERE id = %s", (proveedor_id,))
        nombre_tecnico = (cur.fetchone() or [None])[0] or "No asignado"

        cur.execute("SELECT proceso FROM tecnologia_procesos WHERE id = %s", (proceso_id,))
        proceso_nombre = (cur.fetchone() or [None])[0] or "Sin Proceso"

        cur.execute("SELECT nombre_contratista FROM tecnologia_persona_responsable WHERE id = %s", (persona_id,))
        persona_nombre = (cur.fetchone() or [None])[0] or "No asignado"

        lista_equipos = []
        for r in registros:
            tipo = r.get('tipo')  # fecha_preventivo o fecha_correctivo
            producto_id = r.get('productoId')
            nueva_periodicidad = int(data.get('nuevaPeriodicidad', 0))
            nombre_equipo = r.get('nombreEquipo')
            # ubicacion = r.get('ubicacionOriginal')
            lista_equipos.append({
                "nombre_equipo": nombre_equipo,
                "cod_articulo": producto_id
            })

            # Obtener datos actuales para historial preventivo
            if tipo == "fecha_mantenimiento":
                cur.execute("SELECT fecha_mantenimiento, vencimiento_mantenimiento, periodicidad FROM tecnologia_equipos WHERE cod_articulo = %s", (producto_id,))
                resultado = cur.fetchone()
                if not resultado:
                    continue
                fecha_actual, vencimiento_actual, periodicidad_actual = resultado

                # Guardar en historial preventivo
                cur.execute(
                    """INSERT INTO tecnologia_historial_preventivo 
                    (cod_articulo, nombre_equipo, id_proceso, fecha_mantenimiento, vencimiento_mantenimiento, periodicidad, id_proveedor_responsable, id_persona_responsable, observaciones)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        producto_id,
                        nombre_equipo,
                        proceso_id,
                        fecha_actual,            # fecha anterior guardada
                        nueva_fecha,             # vencimiento será la nueva fecha elegida
                        nueva_periodicidad,
                        proveedor_id,
                        persona_id,
                        observaciones_id,
                    )
                )

                # Calcular nuevo vencimiento
                nuevo_vencimiento = nueva_fecha + relativedelta(months=nueva_periodicidad)

                # Actualizar en tecnologia_equipos solo los preventivos
                cur.execute(
                    "UPDATE tecnologia_equipos SET fecha_mantenimiento = %s, vencimiento_mantenimiento = %s, periodicidad = %s, proveedor_responsable = %s, id_persona_responsable= %s, id_proceso= %s WHERE cod_articulo = %s",
                    (nueva_fecha, nuevo_vencimiento, nueva_periodicidad, proveedor_id, persona_id, proceso_id, producto_id)
                )
            
            # Obtener datos actuales para historial correctivo
            elif tipo == "fecha_calibracion":
                cur.execute("SELECT fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion FROM tecnologia_equipos WHERE cod_articulo = %s", (producto_id,))
                resultado = cur.fetchone()
                if not resultado:
                    continue
                fecha_actual, vencimiento_actual, periodicidad_actual = resultado

                # Guardar en historial correctivo
                cur.execute(
                    """INSERT INTO tecnologia_historial_correctivo 
                    (cod_articulo, nombre_equipo, id_proceso, fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion, id_proveedor_responsable, id_persona_responsable, observaciones)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        producto_id,
                        nombre_equipo,
                        proceso_id,
                        fecha_actual,
                        nueva_fecha,
                        nueva_periodicidad,
                        proveedor_id,
                        persona_id,
                        observaciones_id,
                    )
                )

                # Calcular nuevo vencimiento
                nuevo_vencimiento = nueva_fecha + relativedelta(months=nueva_periodicidad)

                # Actualizar en tecnologia_equipos tanto los correctivos como los preventivos
                cur.execute(
                    "UPDATE tecnologia_equipos SET fecha_calibracion = %s, vencimiento_calibracion = %s, periodicidad_calibracion = %s, fecha_mantenimiento = %s, vencimiento_mantenimiento = %s, periodicidad = %s, proveedor_responsable = %s, id_persona_responsable= %s, id_proceso= %s WHERE cod_articulo = %s",
                    (nueva_fecha, nuevo_vencimiento, nueva_periodicidad, nueva_fecha, nuevo_vencimiento, nueva_periodicidad, proveedor_id, persona_id, proceso_id, producto_id)
                )

        db.connection.commit()

        if lista_equipos:
            send_mantenimiento_notification_html(
                lista_equipos,
                nombre_tecnico=nombre_tecnico,  # puedes obtenerlo con una consulta
                id_proceso=proceso_nombre,  # idem, desde proceso_id
                persona_responsable=persona_nombre,   # desde persona_id
                # observaciones=observaciones_id,
                email_recibe=correo_externo,
                fecha_mantenimiento=nueva_fecha.strftime("%Y-%m-%d"),
                tipo_mantenimiento=tipo
            )
        return jsonify({'success': True, 'message': 'Fechas y registros actualizados correctamente.'})

    except Exception as e:
        db.connection.rollback()
        print("Error:", e)
        return jsonify({'success': False, 'message': f'Error en el servidor: {str(e)}'})
    

# ESTA FUNCIÓN ME LLEVA A OTRA VENTANA TRAYENDO LOS PARAMETROS DE AGREGAR PARA DESPUES PODER ACTUALIZAR EN LA SIGUIENTE FUNCIÓN. LAS DOS SE COMPLEMENTAN
@bp.route('/editEquipoTecnologia/<id>/<vista>', methods=['GET'])
@login_required
def GET_EQUIPO_TECNOLOGIA(id,vista):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Selección de tabla según vista
    if vista == 'indexTecnologia':
        cur.execute("""SELECT id,
           cod_articulo,
           nombre_equipo,
           estado_equipo,
           id_proceso,
           marca_equipo_tecnologia,
           modelo_equipo_tecnologia,
           serial_equipo_tecnologia,
           ram,
           disco,
           tipo_equipo,
           software_instalado,
           fecha_ingreso,
           fecha_de_baja,
           periodicidad,
           fecha_mantenimiento,
           vencimiento_mantenimiento,
           periodicidad_calibracion,
           fecha_calibracion,
           vencimiento_calibracion,
           imagen,
           id_persona_responsable,
           ubicacion
    FROM tecnologia_equipos
    WHERE id = %s""", [id])
    elif vista == 'equiposDeBajaTecnologia':
        cur.execute('SELECT * FROM tecnologia_equipos_debaja WHERE id = %s', [id])

    producto = cur.fetchone()  # ✅ ya es diccionario

    if not producto:
        flash("Equipo no encontrado", "warning")
        return redirect(url_for('indexTecnologia'))

    cod_articulo = producto["cod_articulo"]

    # Historial preventivo
    cur.execute("""SELECT * FROM tecnologia_historial_preventivo 
                   WHERE cod_articulo = %s ORDER BY fecha_mantenimiento DESC""", [cod_articulo])
    historial_mantenimiento = cur.fetchall()

    # Historial correctivo
    cur.execute("""SELECT * FROM tecnologia_historial_correctivo 
                   WHERE cod_articulo = %s ORDER BY fecha_calibracion DESC""", [cod_articulo])
    historial_calibracion = cur.fetchall()

    # Personas responsables → dict id:nombre
    cur.execute('SELECT id, nombre_contratista FROM tecnologia_persona_responsable')
    personas_data = cur.fetchall()
    personas = {p["id"]: p["nombre_contratista"] for p in personas_data}

    # Proceso
    cur.execute('SELECT id, proceso FROM tecnologia_procesos WHERE activo = 1 ORDER BY proceso')
    procesoEquipos_data = cur.fetchall()
    procesoEquipos = {p["id"]: p["proceso"] for p in procesoEquipos_data}

    historial = {
        'preventivo': historial_mantenimiento,
        'correctivo': historial_calibracion
    }
    
    return render_template(
        'editEquipoTecnologia.html',
        producto=producto,
        cod_articulo=cod_articulo,
        historial=historial,
        personas=personas,
        procesoEquipos=procesoEquipos
    )

# FUNCIÓN ACTUALIZAR EDITAR/VER HOJA DE VIDA
@bp.route('/actualizarTecnologia/<id>', methods = ['POST'])
def ACTUALIZAR_EQUIPO_TECNOLOGIA(id):
    if request.method =='POST':
        # cod_articulo = request.form ['cod_articulo']
        nombre_equipo = request.form ['nombre_equipo']
        id_proceso = request.form ['id_proceso'] or None
        ubicacion = request.form ['ubicacion'] or None
        software_instalado = request.form ['software_instalado'] or None
        marca_equipo_tecnologia = request.form ['marca_equipo_tecnologia'] or None
        modelo_equipo_tecnologia = request.form ['modelo_equipo_tecnologia'] or None
        serial_equipo_tecnologia = request.form ['serial_equipo_tecnologia'] or None
        ram = request.form ['ram'] or None
        disco = request.form ['disco'] or None
        
        # PARA EL CHECKBOX Y SEMAFORO DE MANTENIMIENTO
        fecha_mantenimiento = request.form ['fecha_mantenimiento'] or None
        vencimiento_mantenimiento = request.form ['vencimiento_mantenimiento'] or None
        
        # Obtener hora actual del equipo
        hora_actual = datetime.now().date()
        color = 'verde'

        if  vencimiento_mantenimiento:
            vencimiento_mant = datetime.strptime(vencimiento_mantenimiento, '%Y-%m-%d').date()
            if vencimiento_mant < hora_actual:
                color = 'purple'  # Falta menos de un mes
            elif vencimiento_mant <= hora_actual + timedelta(days=30):
                color = 'red'  # Falta menos de tres meses
            elif vencimiento_mant <= hora_actual + timedelta(days=90):
                color = 'yellow'  # Falta menos de tres meses
        
        fecha_calibracion = request.form ['fecha_calibracion'] or None
        vencimiento_calibracion = request.form ['vencimiento_calibracion'] or None

        fecha_ingreso = request.form ['fecha_ingreso'] or None
        periodicidad_raw = request.form ['periodicidad']
        if periodicidad_raw in (None, "", "None"):
            periodicidad = None
        else:
            periodicidad = int(periodicidad_raw)

        periodicidad_calibracion_raw = request.form ['periodicidad_calibracion']
        if periodicidad_calibracion_raw in (None, "", "None"):
            periodicidad_calibracion = None
        else:
            periodicidad_calibracion = int(periodicidad_calibracion_raw)

        # # Ubicacion Original
        # cur.execute('SELECT id, ubicacion_original FROM tecnologia_ubicacion_equipos')
        # ubicacionEquipos_data = cur.fetchall()
        # ubicacionEquipos = {p["id"]: p["ubicacion_original"] for p in ubicacionEquipos_data}
        
        cur = db.connection.cursor() 

        # Obtener las fechas actuales antes de actualizar
        cur.execute(
            """ UPDATE tecnologia_equipos SET  nombre_equipo = %s, id_proceso = %s, ubicacion = %s, software_instalado = %s, marca_equipo_tecnologia = %s, modelo_equipo_tecnologia = %s, 
                serial_equipo_tecnologia =%s, ram =%s, disco =%s, fecha_mantenimiento = %s, vencimiento_mantenimiento = %s, fecha_calibracion = %s, vencimiento_calibracion = %s,
                fecha_ingreso = %s, periodicidad = %s, color = %s, periodicidad_calibracion = %s WHERE id = %s""",
            (
                # cod_articulo,
                nombre_equipo,
                id_proceso,
                ubicacion,
                software_instalado,
                marca_equipo_tecnologia,
                modelo_equipo_tecnologia,
                serial_equipo_tecnologia,
                ram,
                disco,
                fecha_mantenimiento,
                vencimiento_mantenimiento,
                fecha_calibracion,
                vencimiento_calibracion,
                fecha_ingreso,
                periodicidad,
                color,
                periodicidad_calibracion,
                id,
            ),
        )
        db.connection.commit()

        # Obtener datos del equipo
        cur = db.connection.cursor()
        cur.execute("SELECT * FROM tecnologia_equipos WHERE id = %s", (id,))
        producto = cur.fetchone()

        # Obtener todos los procesos
        cur.execute("SELECT id, proceso FROM tecnologia_procesos")
        procesoEquipos = cur.fetchall()

        flash('Equipo actualizado satisfactoriamente', 'success')
        return redirect(url_for('main.indexTecnologia', id=id))
    
# HISTORIAL FECHAS MANTENIMIENTO PREVENTIVO TECNOLOGIA
@bp.route('/historialPreventivoTecnologia/<cod_articulo>')
@login_required
def HISTORIAL_PREVENTIVO_TECNOLOGIA(cod_articulo):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        cur.execute(
            """
            SELECT id, cod_articulo, nombre_equipo, id_proceso, fecha_mantenimiento, 
                   vencimiento_mantenimiento, periodicidad, id_proveedor_responsable, id_persona_responsable, observaciones
            FROM tecnologia_historial_preventivo 
            WHERE cod_articulo = %s 
            ORDER BY fecha_mantenimiento DESC
        """,
            [cod_articulo],
        )
        preventivo = cur.fetchall()

        # Trae el nombre de los técnicos
        cur.execute("SELECT id, nombre_tecnico FROM tecnologia_tecnico_responsable")
        proveedores_data = cur.fetchall()
        proveedores = {p["id"]: p["nombre_tecnico"] for p in proveedores_data}

        # Personas responsables → dict id:nombre
        cur.execute('SELECT id, nombre_contratista FROM tecnologia_persona_responsable')
        personas_data = cur.fetchall()
        personas = {r["id"]: r["nombre_contratista"] for r in personas_data}

        cur.execute('SELECT id, proceso FROM tecnologia_procesos')
        procesoEquipos_data = cur.fetchall()
        procesoEquipos = {p["id"]: p["proceso"] for p in procesoEquipos_data}

        historial = {
            'preventivo': preventivo
            # 'correctivo': correctivo
        }

        return render_template('historialPreventivoTecnologia.html', historial=historial, proveedores=proveedores, personas=personas, procesoEquipos=procesoEquipos)

    except Exception as e:
        print(f"Error al obtener el historial: {str(e)}")
        flash('Error al obtener el historial de fechas.', 'danger')
        return redirect(url_for('indexTecnologia'))
    finally:
        cur.close()


# HISTORIAL FECHAS MANTENIMIENTO CORRECTIVO TECNOLOGIA
@bp.route('/historialCorrectivoTecnologia/<cod_articulo>')
@login_required
def HISTORIAL_CORRECTIVO_TECNOLOGIA(cod_articulo):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        cur.execute(
            """
            SELECT id, cod_articulo, nombre_equipo, id_proceso, fecha_calibracion, 
                   vencimiento_calibracion, periodicidad_calibracion, id_proveedor_responsable, id_persona_responsable, observaciones
            FROM tecnologia_historial_correctivo 
            WHERE cod_articulo = %s 
            ORDER BY fecha_calibracion DESC
        """,
            [cod_articulo],
        )
        correctivo = cur.fetchall()

        # Trae el nombre de los técnicos
        cur.execute("SELECT id, nombre_tecnico FROM tecnologia_tecnico_responsable")
        proveedores_data = cur.fetchall()
        proveedores = {p["id"]: p["nombre_tecnico"] for p in proveedores_data}

        # Personas responsables → dict id:nombre
        cur.execute('SELECT id, nombre_contratista FROM tecnologia_persona_responsable')
        personas_data = cur.fetchall()
        personas = {r["id"]: r["nombre_contratista"] for r in personas_data}

        cur.execute('SELECT id, proceso FROM tecnologia_procesos')
        procesoEquipos_data = cur.fetchall()
        procesoEquipos = {p["id"]: p["proceso"] for p in procesoEquipos_data}

        historial = {
            # 'preventivo': preventivo,
            'correctivo': correctivo
        }

        return render_template('historialCorrectivoTecnologia.html', historial=historial, proveedores=proveedores, personas=personas, procesoEquipos=procesoEquipos)

    except Exception as e:
        print(f"Error al obtener el historial: {str(e)}")
        flash('Error al obtener el historial de fechas.', 'danger')
        return redirect(url_for('indexTecnologia'))
    finally:
        cur.close()


# ACTUALIZAR FECHAS DE MANTENIMIENTO PREVENTIVO
@bp.route('/updateHistorialMantenimientoPreventivo', methods=['POST'])
def update_historial_mantenimiento_preventivo():
    id = request.form['id']  # ID del registro en historial_fechas
    cod_articulo = request.form['cod_articulo']
    fecha_mantenimiento = request.form['fecha_mantenimiento']
    vencimiento_mantenimiento = request.form['vencimiento_mantenimiento']
    periodicidad = request.form['periodicidad']

    # Usar DictCursor para obtener un diccionario
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Verificar si el registro es el último de su tipo en historial_fechas
    cur.execute("""SELECT id FROM tecnologia_historial_preventivo WHERE cod_articulo = %s ORDER BY id DESC LIMIT 1""", [cod_articulo])
    last_record = cur.fetchone()

   # Si el id del último registro coincide con el seleccionado, permitir la actualización
    if last_record and last_record['id'] == int(id):
        # Actualiza
        cur = db.connection.cursor()
        cur.execute("""UPDATE tecnologia_historial_preventivo SET fecha_mantenimiento = %s, 
                       vencimiento_mantenimiento = %s, periodicidad = %s WHERE id = %s""", 
                      (fecha_mantenimiento, vencimiento_mantenimiento, periodicidad, id))
        
        # Determinar el color del semáforo basado en la fecha de vencimiento más próxima
        fecha_actual = datetime.now().date()
        color = "verde"  # Valor por defecto

        # Comparar fechas de mantenimiento y calibración
        if vencimiento_mantenimiento:
            
            vencimiento_mant = datetime.strptime(vencimiento_mantenimiento, '%Y-%m-%d').date()
            if vencimiento_mant < fecha_actual + timedelta(days=0):
                color = "purple"
            elif vencimiento_mant <= fecha_actual + timedelta(days=30):
                color = "red"
            elif vencimiento_mant <= fecha_actual + timedelta(days=90):
                color = "yellow"
        
        db.connection.commit()
        flash('Historial actualizado correctamente', 'success')
    else:
        # Si no es el último registro, mostrar un mensaje de error
        flash('Solo se puede actualizar el último registro de este equipo.', 'danger')
    
    return redirect(url_for('historialPreventivoTecnologia', cod_articulo=cod_articulo))

# ACTUALIZAR FECHAS DE HISTORIAL MENTENIMIENTO CORRECTIVO
@bp.route('/update_historialMantenimientoCorrectivo', methods=['POST'])
def update_historial_mantenimiento_correctivo():
    id = request.form['id']  # ID del registro en historial_fechas_calibracion
    cod_articulo = request.form['cod_articulo']
    fecha_calibracion = request.form['fecha_calibracion']
    vencimiento_calibracion = request.form['vencimiento_calibracion']
    periodicidad_calibracion = request.form['periodicidad_calibracion']

    # Usar DictCursor para obtener un diccionario
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Verificar si el registro es el último de su tipo en historial_fechas
    cur.execute("""SELECT id FROM tecnologia_historial_correctivo WHERE cod_articulo = %s ORDER BY id DESC LIMIT 1 """, [cod_articulo])
    last_record = cur.fetchone()

   # Si el id del último registro coincide con el seleccionado, permitir la actualización
    if last_record and last_record['id'] == int(id):
        # Actualizar
        cur = db.connection.cursor()
        cur.execute(""" UPDATE tecnologia_historial_correctivo SET fecha_calibracion = %s, 
                        vencimiento_calibracion = %s, periodicidad_calibracion = %s WHERE id = %s""", 
                        (fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion, id))
        
        db.connection.commit()
        flash('Historial actualizado correctamente', 'success')
    else:
        # Si no es el último registro, mostrar un mensaje de error
        flash('Solo se puede actualizar el último registro de este equipo.', 'danger')
    
    return redirect(url_for('historialCorrectivoTecnologia', cod_articulo=cod_articulo))

# ======================================================================================================
# ==========================INICIA FUNCIÓN OTROS EQUIPOS DE TECNOLOGIA=====================
@bp.route('/indexOtrosEquiposTecnologia')
@login_required
def index_otros_equipos_tecnologia():
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    # cur = db.connection.cursor()
    cur.execute('SELECT * FROM tecnologia_equipos WHERE enable=0 AND otros_equipos_tecnologia=1')
    data_otros_equipos_tecnologia = cur.fetchall()

    cur.execute('SELECT id, nombre_tecnico FROM tecnologia_tecnico_responsable')
    proveedores = cur.fetchall()

    cur.execute('SELECT id, documento_identidad, nombre_contratista FROM tecnologia_persona_responsable')
    personas = cur.fetchall()

    cur.execute('SELECT id, tipo_equipo FROM tecnologia_tipo_equipo')
    tipoEquipos = cur.fetchall()

    cur.execute('SELECT id, estado_equipo FROM tecnologia_estados_equipos')
    estadoEquipos = cur.fetchall()

    cur.execute('SELECT id, proceso FROM tecnologia_procesos WHERE activo = 1 ORDER BY proceso')
    procesoEquipos = cur.fetchall()

    cur.execute('SELECT id, proceso FROM tecnologia_procesos')
    procesoEquipos_data = cur.fetchall()
    procesoEquiposModal = {p["id"]: p["proceso"] for p in procesoEquipos_data}
    # print(ubicacionEquipos)
    return render_template('IndexOtrosEquiposTecnologia.html', tecnologia_equipos=data_otros_equipos_tecnologia, tipoEquipos=tipoEquipos, proveedores=proveedores, personas=personas, estadoEquipos=estadoEquipos, procesoEquipos=procesoEquipos, procesoEquiposModal=procesoEquiposModal)
# ==========================INICIA FUNCIÓN EQUIPOS DADOS DE BAJA TECNOLOGIA=====================
@bp.route('/equiposDeBajaTecnologia')
@login_required
def equipos_debaja_tecnologia():
    
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute('SELECT id, estado_equipo FROM tecnologia_estados_equipos')
    estadoEquipos = cur.fetchall()
    
    cur.execute('SELECT * FROM tecnologia_equipos_debaja')
    equipos_de_baja_tecnologia = cur.fetchall()

    return render_template('equiposDeBajaTecnologia.html', equipos_de_baja_tecnologia=equipos_de_baja_tecnologia, estadoEquipos=estadoEquipos)
# ==========================FINALIZA FUNCIÓN EQUIPOS DADOS DE BAJA=====================

# FUNCIÓN ELIMINAR
@bp.route('/delete_producto/<string:id>')
def ELIMINAR_CONTACTO(id):
    cur = db.connection.cursor()
    cur.execute('DELETE FROM productos WHERE id = {0}'.format(id))
    db.connection.commit()
    flash('Producto eliminado satisfactoriamente')
    return redirect(url_for('indexTecnologia'))
# --------------------------- FINALIZA MODULO DE TECNOLOGIA-----------------------------

# ---------------------------FUNCIÓN PARA EL MANEJO DE LOS MODULOS-----------------------------
# @bp.route('/<modulo>')
# @login_required
# def index_modulo(modulo):
#     modulos_validos = ['salud', 'gastronomia', 'lacma', 'arquitectura']
    
#     if modulo not in modulos_validos:
#         # flash("Modulo no válido", "error")
#         return redirect(url_for('main.home'))  # <-- redirige al home si no existe
    
#     cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

#     # Traer equipos solo del modulo actual
#     cur.execute("""SELECT i.*, p.enable_prestamos FROM indexssalud i LEFT JOIN prestamos_equiposalud p ON i.cod_articulo = p.cod_articulo AND p.enable_prestamos = 1 WHERE i.enable=1 AND i.de_baja=0 AND i.modulo=%s""", (modulo,))
#     equipos = cur.fetchall()

#     # Traer proveedores, estados y procesos
#     cur.execute('SELECT id, nombre_empresa FROM datosproveedorsalud')
#     proveedores = cur.fetchall()

#     cur.execute('SELECT id, estado_equipo FROM estados_equipos')
#     estadoEquipos = cur.fetchall()

#     cur.execute('SELECT id, ubicacion_original FROM ubicacion_equipos')
#     ubicacionEquipos = cur.fetchall()

#     return render_template(f'indexSalud.html', indexssalud=equipos, proveedores=proveedores, estadoEquipos=estadoEquipos, ubicacionEquipos=ubicacionEquipos, modulo=modulo)

    
# ---------------------------FUNCION PARA CARGAR IMAGEN DEL EQUIPO DESDE LA TABLA indexSalud EN EL CAMPO ACCIONES SUBIR_IMAGEN-----------------------------  
@bp.route('/subir_imagen/<int:id_producto>', methods=['POST'])
def subir_imagen(id_producto):
    if 'imagen_producto' not in request.files:
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(url_for('indexTecnologia'))

    file = request.files['imagen_producto']
    if file.filename == '':
        flash('Por favor seleccione un archivo válido', 'error')
        return redirect(url_for('indexTecnologia'))
    
    # Validar extensión
    if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg')):
        flash('Solo se permiten archivos PNG, JPG', 'error')
        return redirect(url_for('indexTecnologia'))

    if file:
        filename = secure_filename(file.filename)
        filepath_to_db_img = os.path.join('fotos', filename).replace("\\", "/")
        ruta_absoluta = os.path.join(bp.root_path, 'static', filepath_to_db_img)

        # Guardar en disco
        file.save(ruta_absoluta)

        # Actualizar en BD
        cur = db.connection.cursor()
        cur.execute("""
            UPDATE tecnologia_equipos 
            SET imagen = %s 
            WHERE id = %s
        """, (filepath_to_db_img, id_producto))
        db.connection.commit()
        cur.close()

        flash('Imagen cargada correctamente', 'success')
        return redirect(url_for('main.indexTecnologia'))
    
@bp.route('/ver-pdf-baja/<path:filename>')
@login_required
def ver_pdf_baja(filename):
    return send_from_directory(
        UPLOAD_FOLDER_PDF,
        filename,
        as_attachment=False
    )
# ---------------------------INICIA INSERT MASIVO DE EQUIPOS CSV DE SALUD-----------------------------
# @app.route('/insert_csv', methods=['POST'])
# def insert_csv(modulo):
#     if request.method == 'POST':
#         file = request.files['file']
#         if not file:
#             flash('No seleccionó ningún archivo')
#             return redirect(url_for('index_modulo', modulo=modulo))

#         file = TextIOWrapper(file, encoding='latin-1')
#         csv_reader = csv.reader(file)
#         next(csv_reader)  # Saltar encabezado

#         cur = db.connection.cursor()
#         fotos_folder = os.path.join(os.path.dirname(__file__), 'static', 'fotos')

#         codigos_duplicados = []
#         datos_validos = []

#         for row in csv_reader:
#             cod_articulo = row[0]

#             try:
#                 cod_articulo = int(cod_articulo)
#             except ValueError:
#                 flash(f'Código inválido: {cod_articulo}', 'error')
#                 continue  # Salta esta fila

#             cur.execute("SELECT cod_articulo FROM indexssalud WHERE cod_articulo = %s", (cod_articulo,))
#             codigo_indexsalud = cur.fetchone()

#             cur.execute("SELECT cod_articulo FROM equipossalud_debaja WHERE cod_articulo = %s", (cod_articulo,))
#             codigo_debaja = cur.fetchone()

#             if codigo_indexsalud or codigo_debaja:
#                 codigos_duplicados.append(str(cod_articulo))
#                 continue  # Salta esta fila duplicada

#             # Verifica que la imagen exista
#             imagen = row[14]
#             imagen_path = os.path.join(fotos_folder, imagen)
#             if not os.path.isfile(imagen_path):
#                 flash(f'Imagen no encontrada: {imagen}', 'error')
#                 continue  # Salta esta fila

#             # Validación básica de Mantenimiento Actual
#             if not row[7]:
#                 flash(f'Equipo con código {cod_articulo} no tiene Mantenimiento Actual.', 'error')
#                 continue
            
#             # Validación básica de Vencimiento Mantenimiento
#             if not row[8]:
#                 flash(f'Equipo con código {cod_articulo} no tiene Vencimiento Mantenimiento.', 'error')
#                 continue

#             # Validación básica de Periodicidad Calibración
#             if not row[9]:
#                 flash(f'Equipo con código {cod_articulo} si no tiene periodicidad de calibracion, ingresa 0.', 'error')
#                 continue

#             # Preparar fila válida para insertar luego
#             datos_validos.append(row)

#             # Se construye diccionario de proveedores, por nombre de empresa: id
#             cur.execute("SELECT id, nombre_empresa FROM datosproveedorsalud")
#             proveedores = cur.fetchall()
#             proveedor_map = {p[1].strip().lower(): p[0] for p in proveedores}

#         # Insertar solo los datos válidos
#         for row in datos_validos:
#             cod_articulo = int(row[0])
#             nombre_equipo = row[1]
#             ubicacion_original = row[2]
#             estado_equipo = row[3]
#             fecha_ingreso = row[4]
#             garantia = row[5]
#             periodicidad = int(row[6])
#             fecha_mantenimiento = row[7]
#             vencimiento_mantenimiento = row[8]
#             periodicidad_calibracion = int(row[9])
#             fecha_calibracion = row[10] or None
#             vencimiento_calibracion = row[11] or None
#             criticos = row[12]
#             # proveedor_responsable = row[13]
#             nombre_proveedor = row[13].strip().lower()
#             proveedor_responsable = proveedor_map.get(nombre_proveedor)

#             if not proveedor_responsable:
#                 flash(f"Proveedor '{row[13]}' no encontrado en la base de datos.", 'error')
#                 continue
            
#             imagen = row[14]
#             especificaciones_instalacion = row[15]
#             cuidados_basicos = row[16]
#             marca_equipo_salud = row[17]
#             modelo_equipo_salud = row[18]
#             serial_equipo_salud = row[19]

#             ruta_imagen_db = f'fotos/{secure_filename(imagen)}'
#             checkbox_mantenimiento = 'Inactivo'
#             checkbox_calibracion = 'Inactivo'
#             fecha_de_baja = date.today() if estado_equipo == "DE BAJA" else None
#             color = 'verde'

#             if estado_equipo == 'DE BAJA':
#                 # Insertar en la tabla equipossalud_debaja
#                 cur.execute("""INSERT INTO equipossalud_debaja (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, fecha_ingreso, 
#                                                                 periodicidad, estado_equipo, ubicacion_original, garantia, criticos, proveedor_responsable, imagen, especificaciones_instalacion, cuidados_basicos, 
#                                                                 periodicidad_calibracion, marca_equipo_salud, modelo_equipo_salud, serial_equipo_salud, color, checkbox_mantenimiento, checkbox_calibracion, fecha_de_baja) VALUES 
#                                                                 (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
#                     (
#                         cod_articulo,
#                         nombre_equipo,
#                         fecha_mantenimiento,
#                         vencimiento_mantenimiento,
#                         fecha_calibracion,
#                         vencimiento_calibracion,
#                         fecha_ingreso,
#                         periodicidad,
#                         estado_equipo,
#                         ubicacion_original,
#                         garantia,
#                         criticos,
#                         proveedor_responsable,
#                         ruta_imagen_db,
#                         especificaciones_instalacion, 
#                         cuidados_basicos,
#                         periodicidad_calibracion,
#                         marca_equipo_salud,
#                         modelo_equipo_salud,
#                         serial_equipo_salud,
#                         color,
#                         checkbox_mantenimiento,
#                         checkbox_calibracion,
#                         fecha_de_baja,
#                     ),
#                 )
#             else:

#                 # Insertar en la tabla indexssalud
#                 cur.execute("""INSERT INTO indexssalud (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, fecha_ingreso, 
#                                                         periodicidad, estado_equipo, ubicacion_original, garantia, criticos, proveedor_responsable, imagen, especificaciones_instalacion, cuidados_basicos, 
#                                                         periodicidad_calibracion, marca_equipo_salud, modelo_equipo_salud, serial_equipo_salud, color, checkbox_mantenimiento, checkbox_calibracion, fecha_de_baja) VALUES 
#                                                         (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
#                     (
#                         cod_articulo,
#                         nombre_equipo,
#                         fecha_mantenimiento,
#                         vencimiento_mantenimiento,
#                         fecha_calibracion,  
#                         vencimiento_calibracion,  
#                         fecha_ingreso,
#                         periodicidad,
#                         estado_equipo,
#                         ubicacion_original,
#                         garantia,
#                         criticos,
#                         proveedor_responsable,
#                         ruta_imagen_db,
#                         especificaciones_instalacion, 
#                         cuidados_basicos,
#                         periodicidad_calibracion,
#                         marca_equipo_salud,
#                         modelo_equipo_salud,
#                         serial_equipo_salud,
#                         color,
#                         checkbox_mantenimiento,
#                         checkbox_calibracion,
#                         fecha_de_baja,
#                     ),
#                 )

#         db.connection.commit()

#         if codigos_duplicados:
#             flash(f'Los siguientes códigos ya existen y no fueron insertados: {", ".join(codigos_duplicados)}', 'error')
#         if datos_validos:
#             flash(f'{len(datos_validos)} equipos importados exitosamente.', 'success')
#         else:
#             flash('No se importó ningún equipo.', 'error')

#         return redirect(url_for('index_modulo', modulo='modulo'))
# ---------------------------FINALIZA INSERT MASIVO CSV SALUD-----------------------------

# ---------------------------INICIA EXPORTACIÓN DE CSV DE EQUIPOS TECNOLOGIA-----------------------------
@bp.route('/exportCsvTecnologia')
@login_required
def exportCsv():
    cur = db.connection.cursor()

    # Consulta SQL optimizada con JOINs correctos y alias claros
    cur.execute("""
        SELECT 
            i.cod_articulo,
            i.nombre_equipo,
            i.fecha_ingreso,
            i.fecha_mantenimiento,
            i.vencimiento_mantenimiento,
            i.fecha_calibracion,
            i.vencimiento_calibracion,
            i.estado_equipo,
            u.proceso AS id_proceso,
            i.ubicacion,
            i.marca_equipo_tecnologia,
            i.modelo_equipo_tecnologia,
            i.serial_equipo_tecnologia,
            i.ram,
            i.disco,
            i.tipo_equipo,
            i.software_instalado,
            t.nombre_tecnico AS proveedor_responsable,
            q.documento_identidad AS id_persona_responsable,
            p.nombre_contratista AS id_persona_responsable
            # i.enable
        FROM tecnologia_equipos i
        LEFT JOIN tecnologia_tecnico_responsable t ON i.proveedor_responsable = t.id
        LEFT JOIN tecnologia_persona_responsable q ON i.id_persona_responsable = q.id
        LEFT JOIN tecnologia_persona_responsable p ON i.id_persona_responsable = p.id
        LEFT JOIN tecnologia_procesos u ON i.id_proceso = u.id
        WHERE i.de_baja = 0
        ORDER BY i.cod_articulo ASC
    """)

    registros = cur.fetchall()
    cur.close()

    # Crear el archivo CSV en memoria
    si = StringIO()
    writer = csv.writer(si)

    # Encabezados claros y ordenados
    encabezados = [
        'Código Equipo',
        'Nombre Equipo',
        'Fecha Compra',
        'Fecha Ejecución Preventivo',
        'Fecha Vencimiento Preventivo',
        'Fecha Ejecución Correctivo',
        'Fecha Vencimiento Correctivo',
        'Estado Equipo',
        'Proceso',
        'Ubicación',
        'Marca Equipo',
        'Modelo Equipo',
        'Serial Equipo',
        'Memoria RAM',
        'Disco Duro',
        'Tipo Equipo',
        'Software Instalado',
        'Técnico Responsable',
        'ID Responsable del Equipo',
        'Nombre Responsable del Equipo',
        # 'Estado Equipo'
    ]
    writer.writerow(encabezados)

    # Escribir los registros de la tabla
    for registro in registros:
        writer.writerow(registro)

    # Preparar la respuesta de descarga
    salida = Response(
        si.getvalue().encode('utf-8-sig'),
        mimetype='text/csv'
    )
    salida.headers['Content-Disposition'] = 'attachment; filename=equiposTecnologia.csv'

    return salida
# ---------------------------FINALIZA EXPORTACIÓN DE CSV DE EQUIPOS DE SALUD-----------------------------

# ---------------------------INICIA EXPORTACIÓN DE FORMATO EXCEL DE EQUIPOS DE BAJA DE TECNOLIGIA-----------------------------
@bp.route('/exportExcelTecnologiaDeBaja', methods=['POST'])
@login_required
def exportExcelDeBaja():
    try:
        # 1) Validar JSON/csrf
        if not request.is_json:
            return jsonify({"error": "El cuerpo de la petición debe ser JSON"}), 400

        data = request.get_json(silent=True) or {}
        equipos = data.get('equipos') or []

        if not isinstance(equipos, list) or not equipos:
            return jsonify({"error": "No se enviaron equipos"}), 400

        # 2) Ubicar plantilla
        plantilla_path = os.path.join(current_app.root_path, "static", "img", "INFORME_TECNICO_BAJAS.xlsx")
        if not os.path.exists(plantilla_path):
            return jsonify({"error": f"No se encontró la plantilla en {plantilla_path}"}), 400

        # 3) Cargar plantilla
        wb = load_workbook(plantilla_path)
        ws = wb.active  # Ajusta a wb['NombreHoja'] si tu hoja no es la activa

        # 4) Escribir datos desde fila 12 (debajo de los títulos)
        start_row = 12

        # Helper: verificar si un rango ya está fusionado
        def is_merged(ws, coord: str) -> bool:
            return any(str(r) == coord for r in ws.merged_cells.ranges)

        for idx, equipo in enumerate(equipos, start=start_row):
            placa = (equipo.get("cod_articulo") or "").strip()
            cantidad = 1
            nombre = (equipo.get("nombre_equipo") or "").strip()

            # Columna A → PLACA
            ws[f"A{idx}"].value = placa
                       
            # Columna B → CANTIDAD
            ws[f"B{idx}"].value = cantidad

            # Columnas C-G → DESCRIPCIÓN DEL BIEN (escribir en C, y fusionar sólo si hace falta)
            merge_coord = f"C{idx}:G{idx}"
            if not is_merged(ws, merge_coord):
                try:
                    ws.merge_cells(merge_coord)
                except ValueError:
                    # Si ya está fusionado por plantilla o solapa, lo ignoramos
                    pass

            ccell = ws[f"C{idx}"]
            ccell.value = nombre
            ccell.alignment = Alignment(wrap_text=True, vertical="top")

        # 5) Enviar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="equipos_baja.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        current_app.logger.exception("Error generando Excel")
        return jsonify({"error": str(e)}), 400
# ---------------------------FINALIZA EXPORTACIÓN DE CSV DE EQUIPOS DE EQUIPOS DE BAJA DE TECNOLIGIA-----------------------------

# ----------------INICIA FUNCIÓN DE DESCARGUE DE PLANTILLA PARA INSERT MASIVO ----------------------
@bp.route('/download_template_excel')
def download_template_excel_tecnologia():

    wb = Workbook()
    ws = wb.active
    ws.title = "PLANTILLA_TECNOLOGIA"

    # ================================
    # 1) OBTENER PROCESOS DESDE LA BD
    # ================================
    cur = db.connection.cursor()
    cur.execute("SELECT id, proceso FROM tecnologia_procesos WHERE activo = 1 ORDER BY proceso")
    procesos = cur.fetchall()
    cur.close()

    # ================================
    # 2) CREAR HOJA OCULTA DE CATÁLOGOS
    # ================================
    ws_catalogos = wb.create_sheet("CATALOGOS")

    ws_catalogos["A1"] = "PROCESOS"
    ws_catalogos["A1"].font = Font(bold=True)

    for i, p in enumerate(procesos, start=2):
        ws_catalogos.cell(row=i, column=1, value=p[1])  # nombre proceso

    # Ocultar hoja
    ws_catalogos.sheet_state = "hidden"

    total_procesos = len(procesos)

    # ================================
    # 3) ENCABEZADOS (AGREGAMOS PROCESO)
    # ================================
    headers = [
        "Placa Equipo",
        "Nombre Equipo",
        "Fecha de Compra",
        "Tipo de Equipo",
        "Estado de Equipo",
        "Persona Responsable del Equipo",
        "Proceso",
        "Ubicación",
        "Memoria Ram",
        "Disco Duro",
        "Marca Equipo",
        "Modelo Equipo",
        "Serial Equipo",
        "Software Instalado",
        "Ruta Imagen"
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # ================================
    # 4) VALIDACIONES
    # ================================

    # ---- Tipo Equipo
    tipos_equipo = DataValidation(
        type="list",
        formula1='"PORTATIL,TEU (Todo en Uno),CPU,TABLET,VIDEOBEAM"',
        allow_blank=True
    )
    ws.add_data_validation(tipos_equipo)
    tipos_equipo.add("D2:D1000")

    # ---- Estado Equipo
    estados_equipo = DataValidation(
        type="list",
        formula1='"USO,SIN USO,OTROS EQUIPOS"',
        allow_blank=True
    )
    ws.add_data_validation(estados_equipo)
    estados_equipo.add("E2:E1000")

    # ---- Proceso (desde hoja CATALOGOS)
    # Rango dinámico: CATALOGOS!A2:A{N}
    if total_procesos > 0:
        rango_procesos = f"=CATALOGOS!$A$2:$A${total_procesos + 1}"

        procesos_validation = DataValidation(
            type="list",
            formula1=rango_procesos,
            allow_blank=False
        )

        ws.add_data_validation(procesos_validation)
        procesos_validation.add("G2:G1000")  # Columna Proceso

    # ================================
    # 5) ESTILO ENCABEZADOS
    # ================================
    for col in ws[1]:
        col.font = Font(bold=True)
        col.alignment = Alignment(horizontal="center")
        ws.column_dimensions[col.column_letter].width = 28

    # ================================
    # 6) FILA EJEMPLO
    # ================================
    datos = [
        "1",
        "T1",
        "2022/01/01 (Ó CAMPO VACIO)",
        "SELECCIONE UN TIPO",
        "SELECCIONE UN ESTADO",
        "123456(NÚMERO DE DOCTO DEL LIDER DEL PROCESO SIN PUNTOS)",
        "SELECCIONE PROCESO",
        "Auditorio",
        "8GB",
        "500GB",
        "LENOVO THINKPAD",
        "21DBS0NC02",
        "DMPQJ02",
        "Windows 10,Office 2021,7-zip, Adobe, Bitdefender, Google Drive, PowerBI",
        ""
    ]

    for col, value in enumerate(datos, start=1):
        ws.cell(row=2, column=col, value=value)

    # ================================
    # 7) DESCARGAR ARCHIVO
    # ================================
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name="plantilla_carga_inventario_tecnologia.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# ----------------FINALIZA FUNCIÓN DE DESCARGUE DE PLANTILLA PARA INSERT MASIVO ----------------------