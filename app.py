import csv
from io import TextIOWrapper, StringIO, BytesIO

# librerias e importaciones para excel
import io
from flask import request, jsonify, send_file, current_app
from openpyxl import load_workbook
from openpyxl.styles import Alignment
# Importacion del link OneDrive MANTENIEMIENTO desde el archivo config
from config import LinkOneDriveMantenimiento
# Importacion del link OneDrive CALIBRACION desde el archivo config
from config import LinkOneDriveCalibracion

from flask import Flask, render_template, request, redirect, url_for, flash, Response
from flask_mysqldb import MySQL,MySQLdb
from flask_wtf.csrf import CSRFProtect
from flask_login import LoginManager, login_user, logout_user, login_required
from config import config
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta

# Para subir archivos tipo foto al servidor
import os
import shutil
from werkzeug.utils import secure_filename

# Models:
from models.ModelUser import ModelUser

# Entities:
from models.entities.User import User
# Para el modulo json que se esta utilizan para el checked
from flask import Flask, render_template, request, jsonify 
# Importaciones para el envio de correos en prestamo
from flask_login import current_user
# Importaciones desde el archivo email_service
from email_service import send_email_with_logo
from email_service import send_mantenimiento_notification_html
# Importaciones desde el archivo email_devolucion
# from email_devolucion import send_email_envio_with_logo
# from email_devolucion import send_devolucion_notification_html

# app = Flask(__name__, static_url_path='/mantenimientos-tecnologia/static')
app = Flask(
    __name__,
    static_url_path='/mantenimientos-tecnologia/static',
    static_folder='static'
)
app.config.from_object(config['production'])

# Donde configuro mi clave
app.config['SECRET_KEY'] = 'mysecretkey'
app.config['UPLOAD_FOLDER'] = 'static/fotos'

# Inicializo las extenciones
csrf = CSRFProtect(app)
db = MySQL(app)
login_manager_app = LoginManager(app)

@app.context_processor
def link_onedrive_mantenimiento():
    return dict(onedrive_link_mantenimiento=LinkOneDriveMantenimiento.ONEDRIVE_LINK_MANTENIMIENTO)

@app.context_processor
def link_onedrive_calibracion():
    return dict(onedrive_link_calibracion=LinkOneDriveCalibracion.ONEDRIVE_LINK_CALIBRACION)

@login_manager_app.user_loader
def load_user(id):
    return ModelUser.get_by_id(db, id)

@app.after_request
def evita_cache(response):
    response.cache_control.no_store = True
    response.cache_control.no_cache = True
    response.cache_control.must_revalidate = True
    response.cache_control.max_age = 0
    response.expires = 0
    response.pragma = 'no-cache'
    return response

@app.route('/')
# @login_required
def index():
    return redirect(url_for('login'))


@app.route('/mantenimientos-tecnologia/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User(0, request.form['username'], request.form['password'])
        logged_user = ModelUser.login(db, user)

        if logged_user:
            if logged_user.password:
                login_user(logged_user)

                if logged_user.rol in ['salud', 'gastronomia', 'lacma', 'arquitectura', 'tecnologia']:
                    return redirect(url_for('index_modulo', modulo=logged_user.rol))
                elif logged_user.rol == 'admin':
                    return redirect(url_for('home'))
                else:
                    flash('Rol no autorizado')
                    return redirect(url_for('login'))
            else:
                flash("Contraseña incorrecta...")
        else:
            flash("Usuario no encontrado...")
    return render_template('auth/login.html')


@app.route('/mantenimientos-tecnologia/logout')
# @login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/mantenimientos-tecnologia/home')
@login_required
def home():
    cur = db.connection.cursor()

    # Obtener la fecha actual
    fecha_actual = datetime.now().date()

    # Consultar la cantidad de equipos según el vencimiento del mantenimiento
    cur.execute("""
        SELECT 
            SUM(vencimiento_mantenimiento IS NULL) AS sin_fecha,
            SUM(DATEDIFF(vencimiento_mantenimiento, %s) < 0) AS vencidas,
            SUM(DATEDIFF(vencimiento_mantenimiento, %s) BETWEEN 0 AND 30) AS proximas,
            SUM(DATEDIFF(vencimiento_mantenimiento, %s) > 30) AS mas_30_dias
        FROM tecnologia_equipos WHERE enable = 1 and de_baja = 0
    """, (fecha_actual, fecha_actual, fecha_actual))

    # Obtener los resultados
    resultados_preventivos = cur.fetchone()
    sin_fecha_mantenimiento_preventivo = resultados_preventivos[0] or 0
    vencidas_mantenimiento_preventivo = resultados_preventivos[1] or 0
    entre_0_y_30_dias_a_vencer_preventivo = resultados_preventivos[2] or 0
    mas_30_dias_a_vencer_preventivo = resultados_preventivos[3] or 0

    print(resultados_preventivos)
   

    # Consultar la cantidad de equipos según el vencimiento de la calibración
    cur.execute("""
        SELECT 
            SUM(vencimiento_calibracion IS NULL) AS sin_fecha,
            SUM(DATEDIFF(vencimiento_calibracion, %s) < 0) AS vencidas,
            SUM(DATEDIFF(vencimiento_calibracion, %s) BETWEEN 0 AND 30) AS proximas,
            SUM(DATEDIFF(vencimiento_calibracion, %s) > 30) AS mas_30_dias
        FROM tecnologia_equipos WHERE enable = 1 and de_baja = 0
    """, (fecha_actual, fecha_actual, fecha_actual))

    # Obtener los resultados del vencimiento de calibración
    resultados_correctivos = cur.fetchone()
    sin_fecha_mantenimiento_correctivo = resultados_correctivos[0] or 0
    vencidas_mantenimiento_correctivo = resultados_correctivos[1] or 0
    entre_0_y_30_dias_a_vencer_correctivo = resultados_correctivos[2] or 0
    mas_30_dias_a_vencer_correctivo = resultados_correctivos[3] or 0

    return render_template('home.html', sin_fecha_mantenimiento_preventivo=sin_fecha_mantenimiento_preventivo, 
                                        vencidas_mantenimiento_preventivo=vencidas_mantenimiento_preventivo, 
                                        entre_0_y_30_dias_a_vencer_preventivo=entre_0_y_30_dias_a_vencer_preventivo, 
                                        mas_30_dias_a_vencer_preventivo=mas_30_dias_a_vencer_preventivo,
                                        
                                        sin_fecha_mantenimiento_correctivo=sin_fecha_mantenimiento_correctivo, 
                                        vencidas_mantenimiento_correctivo=vencidas_mantenimiento_correctivo, 
                                        entre_0_y_30_dias_a_vencer_correctivo=entre_0_y_30_dias_a_vencer_correctivo, 
                                        mas_30_dias_a_vencer_correctivo=mas_30_dias_a_vencer_correctivo)


# ---------------------------INICIA MODULO DE TECNOLOGIA-----------------------------
@app.route('/datosTecnicoTecnologia/<id>')
@login_required
def DATOS_TECNICO_TECNOLOGIA(id):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    # cur = db.connection.cursor()
    cur.execute('SELECT * FROM tecnologia_tecnico_responsable WHERE id = %s', [id])
    proveedor = cur.fetchall()
    print (proveedor)
    return render_template('datosTecnicoTecnologia.html', tecnologia_tecnico_responsable=proveedor)

# ===============================DATOS PERSONA RESPONSABLE TECNOLOGIA================================
@app.route('/datosPersonaTecnologia/<id>')
@login_required
def DATOS_PERSONA_TECNOLOGIA(id):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    # cur = db.connection.cursor()
    cur.execute('SELECT * FROM tecnologia_persona_responsable WHERE id = %s', [id])
    persona = cur.fetchall()
    print (persona)
    return render_template('datosPersonaTecnologia.html', tecnologia_persona_responsable=persona)

# FUNCIÓN ACTUALIZAR DATOS PERSONAS
@app.route('/update_datos_persona_tecnologia/<id>', methods = ['POST'])
def ACTUALIZAR_DATOS_PERSONA_TECNOLOGIA(id):
    if request.method =='POST':
        documento_identidad = request.form ['documento_identidad']
        # nombre_empresa = request.form ['nombre_empresa']
        nombre_contratista = request.form ['nombre_contratista']
        correo = request.form ['correo']
        # cargo_contacto = request.form ['cargo_contacto']
        area = request.form ['area']
        cur = db.connection.cursor() 
        cur.execute(""" UPDATE tecnologia_persona_responsable SET documento_identidad = %s, nombre_contratista = %s, correo = %s, area = %s WHERE id = %s """, 
                                                      (documento_identidad, nombre_contratista, correo, area, id))
        db.connection.commit()
    flash('Datos actualizados satisfactorimanete', 'success')
    return redirect(url_for('datosPersonaTecnologia', id = id))
    # return redirect(url_for('datosPersonaTecnologia')) 

# ESTA FUNCIÓN ME LLEVA A OTRA VISTA PARA AGREGAR LOS NUEVAS PERSONAS
@app.route('/agregarNuevaPersonaTecnologia')
@login_required
def AGREGAR_NUEVA_PERSONA_TECNOLOGIA():
    return render_template('agregarNuevaPersonaTecnologia.html')

@app.route('/add_datosPersonaTecnologia', methods=['POST'])
def EDITAR_DATOS_PERSONA_TECNOLOGIA():
    if request.method == 'POST':
        documento_identidad = request.form.get('documento_identidad')
        nombre_contratista = request.form.get('nombre_contratista')
        correo = request.form.get('correo')
        area = request.form.get('area')

        # ✅ Valida que todos los campos estén diligenciados
        if not documento_identidad or not nombre_contratista or not correo or not area:
            flash('Todos los campos son obligatorios', 'danger')
            return redirect(url_for('AGREGAR_NUEVA_PERSONA_TECNOLOGIA'))

        cur = db.connection.cursor()

        # ✅ Verifica si ya existe el documento_identidad antes de insertar
        cur.execute(
            "SELECT COUNT(*) FROM tecnologia_persona_responsable WHERE documento_identidad = %s",
            (documento_identidad,)
        )
        existe = cur.fetchone()[0]

        if existe:
            flash('El documento de identidad ya está registrado', 'warning')
            return redirect(url_for('AGREGAR_NUEVA_PERSONA_TECNOLOGIA'))

        # ✅ Inserta solo si no existe
        cur.execute(
            'INSERT INTO tecnologia_persona_responsable (documento_identidad, nombre_contratista, correo, area) '
            'VALUES (%s, %s, %s, %s)',
            (documento_identidad, nombre_contratista, correo, area)
        )
        db.connection.commit()

        flash('Datos agregados satisfactoriamente', 'success')
        return redirect(url_for('AGREGAR_NUEVA_PERSONA_TECNOLOGIA'))
# ============================================================================================

@app.route('/indexTecnologia')
@login_required
def indexTecnologia():
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    # cur = db.connection.cursor()
    cur.execute('SELECT * FROM tecnologia_equipos where enable=1 AND de_baja=0 AND otros_equipos_tecnologia = 0')
    data = cur.fetchall()
    # cur.execute('SELECT i.*, p.enable_prestamos FROM tecnologia_equios i LEFT JOIN prestamos_equiposalud p ON i.cod_articulo = p.cod_articulo AND p.enable_prestamos = 1 WHERE i.enable=1 AND i.de_baja=0') #A raiz del enable=1 no se deben eliminar en la DB
    # data = cur.fetchall()

    cur.execute('SELECT id, nombre_tecnico FROM tecnologia_tecnico_responsable')
    proveedores = cur.fetchall()

    cur.execute('SELECT id, documento_identidad, nombre_contratista FROM tecnologia_persona_responsable')
    personas = cur.fetchall()

    cur.execute('SELECT id, tipo_equipo FROM tecnologia_tipo_equipo')
    tipoEquipos = cur.fetchall()

    cur.execute('SELECT id, estado_equipo FROM tecnologia_estados_equipos')
    estadoEquipos = cur.fetchall()

    cur.execute('SELECT id, ubicacion_original FROM tecnologia_ubicacion_equipos')
    ubicacionEquipos = cur.fetchall()
    
    cur.execute('SELECT id, ubicacion_original FROM tecnologia_ubicacion_equipos')
    ubicacionEquipos_data = cur.fetchall()
    ubicacionEquiposModal = {p["id"]: p["ubicacion_original"] for p in ubicacionEquipos_data}
    
    # print(ubicacionEquipos)
    return render_template('indexTecnologia.html', tecnologia_equipos=data, tipoEquipos=tipoEquipos, proveedores=proveedores, personas=personas, estadoEquipos=estadoEquipos, ubicacionEquipos=ubicacionEquipos, ubicacionEquiposModal=ubicacionEquiposModal)


@app.route('/add_equipos_tecnologia', methods=['POST'])
def AGREGAR_EQUIPOS_TECNOLOGIA():
    if request.method =='POST':
        cod_articulo = request.form ['cod_articulo']
        nombre_equipo = request.form ['nombre_equipo']

        # Validación de cod_articulo
        try:
            cod_articulo = int(cod_articulo)
        except ValueError:
            flash('Por favor ingresar solo números en el código del equipo', 'error')
            return redirect(url_for('indexTecnologia'))

        # Consulta para verificar si el cod_articulo ya existe en la base de datos
        cur = db.connection.cursor()
        cur.execute("SELECT * FROM tecnologia_equipos WHERE cod_articulo = %s", (cod_articulo,))
        existing_articulo = cur.fetchone()

        if existing_articulo:
            flash(f'El código de equipo {cod_articulo} ya existe', 'error')
            return redirect(url_for('indexTecnologia'))

        # PARA EL CHECKBOX Y SEMAFORO DE MANTENIMIENTO
        fecha_mantenimiento = request.form ['fecha_mantenimiento'] or None
        vencimiento_mantenimiento = request.form ['vencimiento_mantenimiento'] or None
        checkbox_mantenimiento = 'Inactivo' # Valor predeterminado
        
        # # Obtener hora actual del equipo
        # hora_actual = datetime.now().date()
        color = 'verde'
        # PARA EL CHECKBOX DE CALIBRACIÓN
        fecha_calibracion = request.form ['fecha_calibracion'] or None
        vencimiento_calibracion = request.form ['vencimiento_calibracion'] or None
        checkbox_calibracion = 'Inactivo' # Valor predeterminado
        fecha_ingreso = request.form ['fecha_ingreso']
        tipo_equipo = request.form ['tipo_equipo']
        estado_equipo = request.form ['estado_equipo']
        ubicacion_original = request.form ['ubicacion_original']
        ram = request.form ['ram']
        disco = request.form ['disco']
        proveedor_responsable = request.form ['proveedor_responsable']

        # Manejo de la imagen
        if 'imagen_producto' not in request.files:
            flash('No existe archivo de imagen.', 'error')
            return redirect(url_for('indexTecnologia'))

        file = request.files['imagen_producto']

        if file.filename == '':
            flash('Por favor seleccione un archivo de imagen.', 'error')
            return redirect(url_for('indexTecnologia'))

        # Validar extensión de imagen
        extensiones_permitidas_img = ('.png', '.jpg', '.jpeg')
        if not file.filename.lower().endswith(extensiones_permitidas_img):
            flash(f'Formato de imagen no permitido. Solo se permiten: {", ".join(extensiones_permitidas_img)}', 'error')
            return redirect(url_for('indexTecnologia'))

        # Guardar imagen
        filename = secure_filename(file.filename)
        filepath_to_db_img = os.path.join('fotos', filename).replace("\\", "/")
        ruta_absoluta = os.path.join(app.root_path, 'static', filepath_to_db_img)
        file.save(ruta_absoluta)

        software_instalado = request.form ['software_instalado']
        marca_equipo_tecnologia = request.form ['marca_equipo_tecnologia']
        modelo_equipo_tecnologia = request.form ['modelo_equipo_tecnologia']
        serial_equipo_tecnologia = request.form ['serial_equipo_tecnologia']
        id_persona_responsable = request.form ['id_persona_responsable']
        otros_equipos_tecnologia = 1 if estado_equipo == "OTROS EQUIPOS" else 0
        fecha_de_baja = date.today() if estado_equipo == "DE BAJA" else None
        
        # Guardar siempre en tecnologia_equipos
        cur.execute("""
            INSERT INTO tecnologia_equipos (
                cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, 
                fecha_calibracion, vencimiento_calibracion, fecha_ingreso, tipo_equipo, 
                estado_equipo, ubicacion_original, ram, disco, proveedor_responsable, 
                color, checkbox_mantenimiento, checkbox_calibracion, imagen, software_instalado, 
                marca_equipo_tecnologia, modelo_equipo_tecnologia, serial_equipo_tecnologia, 
                id_persona_responsable, otros_equipos_tecnologia, de_baja
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
        """, (
            cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento,
            fecha_calibracion, vencimiento_calibracion, fecha_ingreso, tipo_equipo,
            estado_equipo, ubicacion_original, ram, disco, proveedor_responsable,
            color, checkbox_mantenimiento, checkbox_calibracion, filepath_to_db_img, 
            software_instalado, marca_equipo_tecnologia, modelo_equipo_tecnologia,
            serial_equipo_tecnologia, id_persona_responsable, otros_equipos_tecnologia,
            1 if estado_equipo == "DE BAJA" else 0
        ))

        # Si es DE BAJA, guardar también en tecnologia_equipos_debaja
        if estado_equipo == "DE BAJA":
            cur.execute("""
                INSERT INTO tecnologia_equipos_debaja (
                    cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, 
                    fecha_calibracion, vencimiento_calibracion, fecha_ingreso, tipo_equipo,
                    estado_equipo, ubicacion_original, ram, disco, proveedor_responsable, 
                    color, checkbox_mantenimiento, checkbox_calibracion, imagen, 
                    software_instalado, marca_equipo_tecnologia, modelo_equipo_tecnologia, 
                    serial_equipo_tecnologia, fecha_de_baja, id_persona_responsable
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento,
                fecha_calibracion, vencimiento_calibracion, fecha_ingreso, tipo_equipo,
                estado_equipo, ubicacion_original, ram, disco, proveedor_responsable,
                color, checkbox_mantenimiento, checkbox_calibracion, filepath_to_db_img,
                software_instalado, marca_equipo_tecnologia, modelo_equipo_tecnologia,
                serial_equipo_tecnologia, fecha_de_baja, id_persona_responsable
            ))

        db.connection.commit()
        flash("Equipo agregado correctamente", "success")

        # Redirigir según tipo
        if estado_equipo == "OTROS EQUIPOS":
            return redirect(url_for('indexOtrosEquiposTecnologia'))
        else:
            return redirect(url_for('indexTecnologia'))
# ---------------------------INICIA INSERT MASIVO DE EQUIPOS CSV DE TECNOLOGIA-----------------------------
@app.route('/insert_csvTecnologia', methods=['POST'])
def INSERT_CSV_TECNOLOGIA():
    if request.method == 'POST':
        file = request.files['file']
        if not file:
            flash('No seleccionó ningún archivo')
            return redirect(url_for('indexTecnologia'))

        file = TextIOWrapper(file, encoding='latin-1')
        csv_reader = csv.reader(file)
        next(csv_reader)  # Saltar encabezado

        cur = db.connection.cursor()
        fotos_folder = os.path.join(os.path.dirname(__file__), 'static', 'fotos')

        codigos_duplicados = []
        datos_validos = []

        for row in csv_reader:
            cod_articulo = row[0]

            try:
                cod_articulo = int(cod_articulo)
            except ValueError:
                flash(f'Código inválido: {cod_articulo}', 'error')
                continue  # Salta esta fila

            cur.execute("SELECT cod_articulo FROM tecnologia_equipos WHERE cod_articulo = %s", (cod_articulo,))
            codigo_indextecnologia = cur.fetchone()

            cur.execute("SELECT cod_articulo FROM tecnologia_equipos_debaja WHERE cod_articulo = %s", (cod_articulo,))
            codigo_debaja_tecnologia = cur.fetchone()

            if codigo_indextecnologia or codigo_debaja_tecnologia:
                codigos_duplicados.append(str(cod_articulo))
                continue  # Salta esta fila duplicada

            # Verifica que la imagen exista
            imagen = row[19]
            imagen_path = os.path.join(fotos_folder, imagen)
            if not os.path.isfile(imagen_path):
                flash(f'Imagen no encontrada: {imagen}', 'error')
                continue  # Salta esta fila

            # Preparar fila válida para insertar luego
            datos_validos.append(row)

            # Se construye diccionario de tecnicos, por nombre de tecnicos: id
            cur.execute("SELECT id, nombre_tecnico FROM tecnologia_tecnico_responsable")
            proveedores = cur.fetchall()
            proveedor_map = {p[1].strip().lower(): p[0] for p in proveedores}

            # Se construye diccionario de personas, por nombre de documento de identidad: id
            cur.execute("SELECT id, documento_identidad FROM tecnologia_persona_responsable")
            personas = cur.fetchall()
            persona_map = {str(p[1]).strip().lower(): p[0] for p in personas}

        # Insertar solo los datos válidos
        for row in datos_validos:
            cod_articulo = int(row[0])
            nombre_equipo = row[1]
            fecha_ingreso = row[2]
            periodicidad = int(row[3]) if row[3].strip() else None
            fecha_mantenimiento = row[4] or None
            vencimiento_mantenimiento = row[5] or None
            periodicidad_calibracion = int(row[6]) if row[6].strip() else None
            fecha_calibracion = row[7] or None
            vencimiento_calibracion = row[8] or None
            tipo_equipo = row[9] or None
            estado_equipo = row[10]
            # ubicacion_original = row[11]
            nombre_tecnico = row[11].strip().lower()
            proveedor_responsable = proveedor_map.get(nombre_tecnico)
            nombre_persona = row[12].strip().lower()
            id_persona_responsable = persona_map.get(nombre_persona) 

            if not id_persona_responsable:
                flash(f"Persona '{row[12]}' no encontrado en la base de datos.", 'error')
                continue
            
            ram = row[13]
            disco = row[14]
            marca_equipo_tecnologia = row[15]
            modelo_equipo_tecnologia = row[16]
            serial_equipo_tecnologia = row[17]
            software_instalado = row [18]
            imagen = row[19]


            ruta_imagen_db = f'fotos/{secure_filename(imagen)}'
            checkbox_mantenimiento = 'Inactivo'
            checkbox_calibracion = 'Inactivo'
            fecha_de_baja = date.today() if estado_equipo == "DE BAJA" else None
            color = 'verde'

            if estado_equipo == 'DE BAJA':
                # Insertar en la tabla tecnologia_equipos_debaja
                cur.execute("""INSERT INTO tecnologia_equipos_debaja (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, fecha_ingreso, tipo_equipo,
                           estado_equipo, ram, disco, proveedor_responsable, color, checkbox_mantenimiento, checkbox_calibracion, imagen, software_instalado, marca_equipo_tecnologia, modelo_equipo_tecnologia, serial_equipo_tecnologia, fecha_de_baja, id_persona_responsable) 
                        VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (
                    cod_articulo,
                    nombre_equipo,
                    fecha_mantenimiento,
                    vencimiento_mantenimiento,
                    fecha_calibracion,
                    vencimiento_calibracion,
                    fecha_ingreso,
                    # periodicidad,
                    tipo_equipo,
                    estado_equipo,
                    # ubicacion_original,
                    ram,
                    disco,
                    proveedor_responsable,
                    color,
                    checkbox_mantenimiento,
                    checkbox_calibracion,
                    ruta_imagen_db,
                    software_instalado,
                    # cuidados_basicos,
                    # periodicidad_calibracion,
                    marca_equipo_tecnologia,
                    modelo_equipo_tecnologia,
                    serial_equipo_tecnologia,
                    fecha_de_baja,
                    id_persona_responsable,
                    # color,
                    # checkbox_mantenimiento,
                    # checkbox_calibracion,
                    # fecha_de_baja,
                    
                ),
            )
            else:

                # Insertar en la tabla tecnologia_equipos
                cur.execute("""INSERT INTO tecnologia_equipos (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, fecha_ingreso, tipo_equipo,
                           estado_equipo, ram, disco, proveedor_responsable, color, checkbox_mantenimiento, checkbox_calibracion, imagen, software_instalado, marca_equipo_tecnologia, modelo_equipo_tecnologia, serial_equipo_tecnologia, id_persona_responsable) 
                        VALUES (  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (
                    cod_articulo,
                    nombre_equipo,
                    fecha_mantenimiento,
                    vencimiento_mantenimiento,
                    fecha_calibracion,
                    vencimiento_calibracion,
                    fecha_ingreso,
                    # periodicidad,
                    tipo_equipo,
                    estado_equipo,
                    # ubicacion_original,
                    ram,
                    disco,
                    proveedor_responsable,
                    color,
                    checkbox_mantenimiento,
                    checkbox_calibracion,
                    ruta_imagen_db,
                    software_instalado,
                    # cuidados_basicos,
                    # periodicidad_calibracion,
                    marca_equipo_tecnologia,
                    modelo_equipo_tecnologia,
                    serial_equipo_tecnologia,
                    id_persona_responsable,
                    # color,
                    # checkbox_mantenimiento,
                    # checkbox_calibracion,
                    # fecha_de_baja,

                ),
            )

        db.connection.commit()

        if codigos_duplicados:
            flash(f'Los siguientes códigos ya existen y no fueron insertados: {", ".join(codigos_duplicados)}', 'error')
        if datos_validos:
            flash(f'{len(datos_validos)} equipos importados exitosamente.', 'success')
        else:
            flash('No se importó ningún equipo.', 'error')

        return redirect(url_for('index_modulo', modulo='modulo'))
# ---------------------------FINALIZA INSERT MASIVO CSV DE SALUD-----------------------------
    
# ================================CHECKBOX PROGRAMACIÓN MANTENIMIENTO TECNOLOGIA===============================
@app.route('/checkbox_programacionMantenimientoTecnologia', methods=['POST'])
def checkbox_programacion_mantenimiento_tecnologia():
    try:
        seleccionados = request.form.getlist('seleccionados[]')
        proveedor_id = request.form.get('proveedor_id')
        persona_id = request.form.get('persona_id')
        ubicacion_id = request.form.get('ubicacion_id')

        if not seleccionados or not proveedor_id:
            return jsonify({'success': False, 'message': 'Faltan productos seleccionados o proveedor.'})

        hoy = datetime.now()
        cur = db.connection.cursor()

        productos_guardados = []

        for cod in seleccionados:
            nombre_equipo = request.form.get(f'nombre_equipo_{cod}')
            # ubicacion_id = request.form.get(f'ubicacion_{cod}')
            periodicidad_m = request.form.get(f'periodicidad_mantenimiento_{cod}')
            periodicidad_c = request.form.get(f'periodicidad_calibracion_{cod}')

            mantenimiento_activado = request.form.get(f'mantenimiento_{cod}') == 'on'
            calibracion_activada = request.form.get(f'calibracion_{cod}') == 'on'

            # Obtener fechas reales desde base de datos
            cur.execute("SELECT fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion FROM tecnologia_equipos WHERE cod_articulo = %s", (cod,))
            resultado = cur.fetchone()
            if not resultado:
                continue  # Saltar si no se encuentra

            fecha_m, vencimiento_m, fecha_c, vencimiento_c = resultado

            if mantenimiento_activado:
                # Validación de vencimiento
                if vencimiento_m and (vencimiento_m - hoy).days < 30:
                    continue
                # Activar checkbox y guardar en historial preventivo
                cur.execute(
                    "UPDATE tecnologia_equipos SET checkbox_mantenimiento = 'Activo' WHERE cod_articulo = %s",
                    (cod,),
                )
                cur.execute(
                    """INSERT INTO tecnologia_historial_preventivo 
                            (cod_articulo, nombre_equipo, ubicacion_original, fecha_mantenimiento, vencimiento_mantenimiento, periodicidad, id_proveedor_responsable, id_persona_responsable) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        cod,
                        nombre_equipo,
                        ubicacion_id,
                        fecha_m,
                        vencimiento_m,
                        periodicidad_m,
                        proveedor_id,
                        persona_id,

                    ),
                )

            if calibracion_activada:
                if vencimiento_c and (vencimiento_c - hoy).days < 30:
                    continue
                # Activar checkbox y guardar en historial correctivo
                cur.execute(
                    "UPDATE tecnologia_equipos SET checkbox_calibracion = 'Activo' WHERE cod_articulo = %s",
                    (cod,),
                )
                cur.execute(
                    """INSERT INTO tecnologia_historial_correctivo 
                            (cod_articulo, nombre_equipo, ubicacion_original, fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion, id_proveedor_responsable, id_persona_responsable) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        cod,
                        nombre_equipo,
                        ubicacion_id,
                        fecha_c,
                        vencimiento_c,
                        periodicidad_c,
                        proveedor_id,
                        persona_id,
                    ),
                )

            productos_guardados.append(cod)

        db.connection.commit()
        return jsonify({
            'success': True,
            'message': f'Se procesaron {len(productos_guardados)} productos correctamente.',
            'productos': productos_guardados
        })

    except Exception as e:
        db.connection.rollback()
        return jsonify({'success': False, 'message': f'Error en el servidor: {str(e)}'})

# OBTIENE LOS DATOS DEL RESPONSABLE DEL EQUIPO Y LLENA LOS CAMPOS DE PERSONA RESPONSABLE Y PROCESO DEL MODAL
@app.route('/get_datos_persona/<id>', methods=['GET'])
def get_datos_persona(id):
    try:
        cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

        cur.execute("""
            SELECT 
                p.id AS persona_id,
                p.nombre_contratista,
                u.id AS ubicacion_id,
                u.ubicacion_original
            FROM tecnologia_equipos e
            LEFT JOIN tecnologia_persona_responsable p 
                ON e.id_persona_responsable = p.id
            LEFT JOIN tecnologia_ubicacion_equipos u 
                ON e.ubicacion_original = u.id
            WHERE e.cod_articulo = %s
            LIMIT 1
        """, (id,))

        data = cur.fetchone()
        cur.close()

        if data:
            return jsonify({'success': True, **data})
        else:
            return jsonify({'success': False, 'message': 'No se encontraron datos.'})

    except Exception as e:
        print(f"⚠️ Error en get_datos_persona: {e}")
        return jsonify({'success': False, 'message': 'Error al obtener datos.'})

    
# ACTUALIZA EL ESTADO DEL EQUIPO DESDE EL DESPLEGABLE QUE SE ENCUENTRA EN LA MISMA TABLA INDEXSALUD
@app.route('/update_estadoEquipoTecnologia', methods=['POST'])
def update_estado_equipo_tecnologia():
    if request.method == 'POST':

        # OBTENER FULLNAME DEL USUARIO LOGUEADO
        cur = db.connection.cursor()
        cur.execute("SELECT fullname, username FROM user WHERE id = %s", (current_user.id,))
        result = cur.fetchone()
        usuario_logueado_nombre = result[0] if result else None
        usuario_logueado_email  = result[1] if result else None

        producto_id = request.form['producto_id']
        nuevo_estado = request.form['nuevo_estado_equipo']
        cod_articulo = request.form ['cod_articulo']
        nombre_equipo = request.form ['nombre_equipo']

        # Obtener hora actual del equipo
        hora_actual = datetime.now()

        # PARA EL CHECKBOX Y SEMAFORO DE MANTENIMIENTO
        fecha_mantenimiento = request.form ['fecha_mantenimiento']
        vencimiento_mantenimiento = request.form ['vencimiento_mantenimiento']
        color = 'verde'
        
        # PARA EL CHECKBOX DE CALIBRACIÓN
        fecha_calibracion = request.form.get('fecha_calibracion', '') or None
        vencimiento_calibracion = request.form.get('vencimiento_calibracion', '') or None
       
        # Convertir a objetos date solo si existen
        if fecha_calibracion:
            try:
                fecha_calibracion = datetime.strptime(fecha_calibracion, '%Y-%m-%d').date()
            except ValueError:
                fecha_calibracion = None

        if vencimiento_calibracion:
            try:
                vencimiento_calibracion = datetime.strptime(vencimiento_calibracion, '%Y-%m-%d').date()
            except ValueError:
                vencimiento_calibracion = None

        fecha_ingreso = request.form ['fecha_ingreso']
        periodicidad = request.form ['periodicidad']
        tipo_equipo = request.form ['tipo_equipo']
        estado_equipo= request.form ['estado_equipo']
        ubicacion_original = request.form ['ubicacion_original']
        ram = request.form ['ram']
        disco = request.form ['disco']
        proveedor_responsable = request.form ['proveedor_responsable']
        software_instalado = request.form ['software_instalado']
        # cuidados_basicos = request.form ['cuidados_basicos']
        periodicidad_calibracion = request.form ['periodicidad_calibracion']
        marca_equipo_tecnologia = request.form ['marca_equipo_tecnologia']
        modelo_equipo_tecnologia = request.form ['modelo_equipo_tecnologia']
        serial_equipo_tecnologia = request.form ['serial_equipo_tecnologia']
        id_persona_responsable = request.form ['id_persona_responsable']
        cur = db.connection.cursor()

        # Obtener la ruta de la imagen desde la tabla tecnologia_equipos
        cur.execute('SELECT imagen FROM tecnologia_equipos WHERE cod_articulo = %s', (cod_articulo,))
        imagen_result = cur.fetchone()
        filepath_to_db_img = imagen_result[0] if imagen_result else None

        if nuevo_estado == 'DE BAJA':
            # Actualizar el estado y marcar como dado de baja en tecnologia_equipos
            cur.execute("""UPDATE tecnologia_equipos SET estado_equipo = %s, enable = 0, de_baja = 1, otros_equipos_tecnologia = 0 WHERE cod_articulo = %s""", (nuevo_estado, cod_articulo))

            # Verificar si el equipo ya está en tecnologia_equipos_debaja
            cur.execute('SELECT 1 FROM tecnologia_equipos_debaja WHERE cod_articulo = %s', (cod_articulo,))
            equipo_existente = cur.fetchone()

            # Insertar el equipo en tecnologia_equipos_debaja si no existe
            if not equipo_existente:
                cur.execute("""INSERT INTO tecnologia_equipos_debaja (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, fecha_ingreso,
                                                                periodicidad, tipo_equipo, estado_equipo, ubicacion_original, ram, disco, proveedor_responsable, color, imagen, software_instalado,
                                                                periodicidad_calibracion, marca_equipo_tecnologia, modelo_equipo_tecnologia, serial_equipo_tecnologia, id_persona_responsable, fecha_de_baja) 
                                                                VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        cod_articulo,
                        nombre_equipo,
                        fecha_mantenimiento,
                        vencimiento_mantenimiento,
                        fecha_calibracion,
                        vencimiento_calibracion,
                        fecha_ingreso,
                        periodicidad,
                        tipo_equipo,
                        nuevo_estado,  # Estado del equipo
                        ubicacion_original,
                        ram,
                        disco,
                        proveedor_responsable,
                        color,
                        filepath_to_db_img,
                        software_instalado,
                        # cuidados_basicos,
                        periodicidad_calibracion,
                        marca_equipo_tecnologia,
                        modelo_equipo_tecnologia,
                        serial_equipo_tecnologia,
                        id_persona_responsable,
                        hora_actual
                    ),
                )
        
        # --- ESTADO: OTROS EQUIPOS ---
        elif nuevo_estado == 'OTROS EQUIPOS':
            cur.execute("""
                UPDATE tecnologia_equipos 
                SET estado_equipo = %s, otros_equipos_tecnologia = 1, enable = 0, de_baja = 0 
                WHERE cod_articulo = %s
            """, (nuevo_estado, cod_articulo))

        else:

        # --- 2. Actualizar el estado del equipo ---
            cur.execute("""
                UPDATE tecnologia_equipos
                SET estado_equipo = %s, otros_equipos_tecnologia = 0 
                WHERE cod_articulo = %s
            """, (nuevo_estado, cod_articulo))

        db.connection.commit()
        cur.close()
        flash('Estado del equipo actualizado correctamente', 'success')
        return redirect(url_for('indexTecnologia'))
# =========================================================================================================    

@app.route('/guardar_historialTecnologia', methods=['POST'])
@login_required
def guardar_historial_tecnologia():
    data = request.get_json()
    proveedor_id = data.get('proveedorId')
    persona_id = data.get('personaId')
    ubicacion_id = data.get('ubicacionId')
    observaciones_id = data.get('observacionesId')
    nueva_fecha_str = data.get('nuevaFecha')
    correo_externo = data.get('correoExterno')
    registros = data.get('registros', [])

    try:
        if not proveedor_id or not nueva_fecha_str:
            return jsonify({'success': False, 'message': 'Falta proveedor o fecha'})

        nueva_fecha = datetime.strptime(nueva_fecha_str, '%Y-%m-%d')
        cur = db.connection.cursor()

        # --- Obtener nombres descriptivos ---
        cur.execute("SELECT nombre_tecnico FROM tecnologia_tecnico_responsable WHERE id = %s", (proveedor_id,))
        nombre_tecnico = (cur.fetchone() or [None])[0] or "No asignado"

        cur.execute("SELECT ubicacion_original FROM tecnologia_ubicacion_equipos WHERE id = %s", (ubicacion_id,))
        ubicacion_nombre = (cur.fetchone() or [None])[0] or "Sin ubicación"

        cur.execute("SELECT nombre_contratista FROM tecnologia_persona_responsable WHERE id = %s", (persona_id,))
        persona_nombre = (cur.fetchone() or [None])[0] or "No asignado"


        for r in registros:
            tipo = r.get('tipo')  # fecha_preventivo o fecha_correctivo
            producto_id = r.get('productoId')
            nueva_periodicidad = int(data.get('nuevaPeriodicidad', 0))
            nombre_equipo = r.get('nombreEquipo')
            # ubicacion = r.get('ubicacionOriginal')

            # Obtener datos actuales para historial preventivo
            if tipo == "fecha_mantenimiento":
                cur.execute("SELECT fecha_mantenimiento, vencimiento_mantenimiento, periodicidad FROM tecnologia_equipos WHERE cod_articulo = %s", (producto_id,))
                resultado = cur.fetchone()
                if not resultado:
                    continue
                fecha_actual, vencimiento_actual, periodicidad_actual = resultado

                # Guardar en historial preventivo
                cur.execute(
                    """INSERT INTO tecnologia_historial_preventivo 
                    (cod_articulo, nombre_equipo, ubicacion_original, fecha_mantenimiento, vencimiento_mantenimiento, periodicidad, id_proveedor_responsable, id_persona_responsable, observaciones)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        producto_id,
                        nombre_equipo,
                        ubicacion_id,
                        fecha_actual,            # fecha anterior guardada
                        nueva_fecha,             # vencimiento será la nueva fecha elegida
                        nueva_periodicidad,
                        proveedor_id,
                        persona_id,
                        observaciones_id,
                    )
                )

                # Calcular nuevo vencimiento
                nuevo_vencimiento = nueva_fecha + relativedelta(months=nueva_periodicidad)

                # Actualizar en tecnologia_equipos solo los preventivos
                cur.execute(
                    "UPDATE tecnologia_equipos SET fecha_mantenimiento = %s, vencimiento_mantenimiento = %s, periodicidad = %s, proveedor_responsable = %s, id_persona_responsable= %s, ubicacion_original= %s WHERE cod_articulo = %s",
                    (nueva_fecha, nuevo_vencimiento, nueva_periodicidad, proveedor_id, persona_id, ubicacion_id, producto_id)
                )
            
            # Obtener datos actuales para historial correctivo
            elif tipo == "fecha_calibracion":
                cur.execute("SELECT fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion FROM tecnologia_equipos WHERE cod_articulo = %s", (producto_id,))
                resultado = cur.fetchone()
                if not resultado:
                    continue
                fecha_actual, vencimiento_actual, periodicidad_actual = resultado

                # Guardar en historial correctivo
                cur.execute(
                    """INSERT INTO tecnologia_historial_correctivo 
                    (cod_articulo, nombre_equipo, ubicacion_original, fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion, id_proveedor_responsable, id_persona_responsable, observaciones)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        producto_id,
                        nombre_equipo,
                        ubicacion_id,
                        fecha_actual,
                        nueva_fecha,
                        nueva_periodicidad,
                        proveedor_id,
                        persona_id,
                        observaciones_id,
                    )
                )

                # Calcular nuevo vencimiento
                nuevo_vencimiento = nueva_fecha + relativedelta(months=nueva_periodicidad)

                # Actualizar en tecnologia_equipos tanto los correctivos como los preventivos
                cur.execute(
                    "UPDATE tecnologia_equipos SET fecha_calibracion = %s, vencimiento_calibracion = %s, periodicidad_calibracion = %s, fecha_mantenimiento = %s, vencimiento_mantenimiento = %s, periodicidad = %s, proveedor_responsable = %s, id_persona_responsable= %s, ubicacion_original= %s WHERE cod_articulo = %s",
                    (nueva_fecha, nuevo_vencimiento, nueva_periodicidad, nueva_fecha, nuevo_vencimiento, nueva_periodicidad, proveedor_id, persona_id, ubicacion_id, producto_id)
                )

        db.connection.commit()

        send_mantenimiento_notification_html(
            nombre_equipo=nombre_equipo,
            cod_articulo=producto_id,
            nombre_tecnico=nombre_tecnico,  # puedes obtenerlo con una consulta
            ubicacion_original=ubicacion_nombre,  # idem, desde ubicacion_id
            persona_responsable=persona_nombre,   # desde persona_id
            # observaciones=observaciones_id,
            email_recibe=correo_externo,
            fecha_mantenimiento=nueva_fecha.strftime("%Y-%m-%d"),
            tipo_mantenimiento=tipo
        )
        return jsonify({'success': True, 'message': 'Fechas y registros actualizados correctamente.'})

    except Exception as e:
        db.connection.rollback()
        print("Error:", e)
        return jsonify({'success': False, 'message': f'Error en el servidor: {str(e)}'})
    

# ESTA FUNCIÓN ME LLEVA A OTRA VENTANA TRAYENDO LOS PARAMETROS DE AGREGAR PARA DESPUES PODER ACTUALIZAR EN LA SIGUIENTE FUNCIÓN. LAS DOS SE COMPLEMENTAN
@app.route('/editEquipoTecnologia/<id>/<vista>', methods=['GET'])
@login_required
def GET_EQUIPO_TECNOLOGIA(id,vista):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Selección de tabla según vista
    if vista == 'indexTecnologia':
        cur.execute("""SELECT id,
           cod_articulo,
           nombre_equipo,
           estado_equipo,
           ubicacion_original,
           marca_equipo_tecnologia,
           modelo_equipo_tecnologia,
           serial_equipo_tecnologia,
           ram,
           disco,
           tipo_equipo,
           software_instalado,
           fecha_ingreso,
           fecha_de_baja,
           periodicidad,
           fecha_mantenimiento,
           vencimiento_mantenimiento,
           periodicidad_calibracion,
           fecha_calibracion,
           vencimiento_calibracion,
           imagen,
           id_persona_responsable
    FROM tecnologia_equipos
    WHERE id = %s""", [id])
    elif vista == 'equiposDeBajaTecnologia':
        cur.execute('SELECT * FROM tecnologia_equipos_debaja WHERE id = %s', [id])

    producto = cur.fetchone()  # ✅ ya es diccionario

    if not producto:
        flash("Equipo no encontrado", "warning")
        return redirect(url_for('indexTecnologia'))

    cod_articulo = producto["cod_articulo"]

    # Historial preventivo
    cur.execute("""SELECT * FROM tecnologia_historial_preventivo 
                   WHERE cod_articulo = %s ORDER BY fecha_mantenimiento DESC""", [cod_articulo])
    historial_mantenimiento = cur.fetchall()

    # Historial correctivo
    cur.execute("""SELECT * FROM tecnologia_historial_correctivo 
                   WHERE cod_articulo = %s ORDER BY fecha_calibracion DESC""", [cod_articulo])
    historial_calibracion = cur.fetchall()

    # Personas responsables → dict id:nombre
    cur.execute('SELECT id, nombre_contratista FROM tecnologia_persona_responsable')
    personas_data = cur.fetchall()
    personas = {p["id"]: p["nombre_contratista"] for p in personas_data}

    # Ubicacion Original
    cur.execute('SELECT id, ubicacion_original FROM tecnologia_ubicacion_equipos')
    ubicacionEquipos_data = cur.fetchall()
    ubicacionEquipos = {p["id"]: p["ubicacion_original"] for p in ubicacionEquipos_data}

    historial = {
        'preventivo': historial_mantenimiento,
        'correctivo': historial_calibracion
    }
    
    return render_template(
        'editEquipoTecnologia.html',
        producto=producto,
        cod_articulo=cod_articulo,
        historial=historial,
        personas=personas,
        ubicacionEquipos=ubicacionEquipos
    )

# FUNCIÓN ACTUALIZAR EDITAR/VER HOJA DE VIDA
@app.route('/actualizarTecnologia/<id>', methods = ['POST'])
def ACTUALIZAR_EQUIPO_TECNOLOGIA(id):
    if request.method =='POST':
        cod_articulo = request.form ['cod_articulo']
        nombre_equipo = request.form ['nombre_equipo']
        ubicacion_original = request.form ['ubicacion_original'] or None
        
        # PARA EL CHECKBOX Y SEMAFORO DE MANTENIMIENTO
        fecha_mantenimiento = request.form ['fecha_mantenimiento'] or None
        vencimiento_mantenimiento = request.form ['vencimiento_mantenimiento'] or None
        
        # Obtener hora actual del equipo
        hora_actual = datetime.now().date()
        color = 'verde'

        if  vencimiento_mantenimiento:
            vencimiento_mant = datetime.strptime(vencimiento_mantenimiento, '%Y-%m-%d').date()
            if vencimiento_mant < hora_actual:
                color = 'purple'  # Falta menos de un mes
            elif vencimiento_mant <= hora_actual + timedelta(days=30):
                color = 'red'  # Falta menos de tres meses
            elif vencimiento_mant <= hora_actual + timedelta(days=90):
                color = 'yellow'  # Falta menos de tres meses
        
        fecha_calibracion = request.form ['fecha_calibracion'] or None
        vencimiento_calibracion = request.form ['vencimiento_calibracion'] or None

        fecha_ingreso = request.form ['fecha_ingreso']
        periodicidad = request.form ['periodicidad']
        periodicidad_calibracion = request.form ['periodicidad_calibracion']

        # # Ubicacion Original
        # cur.execute('SELECT id, ubicacion_original FROM tecnologia_ubicacion_equipos')
        # ubicacionEquipos_data = cur.fetchall()
        # ubicacionEquipos = {p["id"]: p["ubicacion_original"] for p in ubicacionEquipos_data}
        
        cur = db.connection.cursor() 

        # Obtener las fechas actuales antes de actualizar
        cur.execute(
            """ UPDATE tecnologia_equipos SET cod_articulo = %s, nombre_equipo = %s, ubicacion_original = %s, fecha_mantenimiento = %s, vencimiento_mantenimiento = %s, fecha_calibracion = %s, vencimiento_calibracion = %s,
                fecha_ingreso = %s, periodicidad = %s, color = %s, periodicidad_calibracion = %s WHERE id = %s""",
            (
                cod_articulo,
                nombre_equipo,
                ubicacion_original,
                fecha_mantenimiento,
                vencimiento_mantenimiento,
                fecha_calibracion,
                vencimiento_calibracion,
                fecha_ingreso,
                periodicidad,
                color,
                # software_instalado,
                # cuidados_basicos,
                periodicidad_calibracion,
                
                id,
            ),
        )
        db.connection.commit()

        # Obtener datos del equipo
        cur = db.connection.cursor()
        cur.execute("SELECT * FROM tecnologia_equipos WHERE id = %s", (id,))
        producto = cur.fetchone()

        # Obtener todas las ubicaciones
        cur.execute("SELECT id, ubicacion_original FROM tecnologia_ubicacion_equipos")
        ubicacionEquipos = cur.fetchall()

        flash('Equipo actualizado satisfactoriamente', 'success')
        return redirect(url_for('indexTecnologia', id=id))
    
# HISTORIAL FECHAS MANTENIMIENTO PREVENTIVO TECNOLOGIA
@app.route('/historialPreventivoTecnologia/<cod_articulo>')
@login_required
def HISTORIAL_PREVENTIVO_TECNOLOGIA(cod_articulo):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        cur.execute(
            """
            SELECT id, cod_articulo, nombre_equipo, ubicacion_original, fecha_mantenimiento, 
                   vencimiento_mantenimiento, periodicidad, id_proveedor_responsable, id_persona_responsable, observaciones
            FROM tecnologia_historial_preventivo 
            WHERE cod_articulo = %s 
            ORDER BY fecha_mantenimiento DESC
        """,
            [cod_articulo],
        )
        preventivo = cur.fetchall()

        # Trae el nombre de los técnicos
        cur.execute("SELECT id, nombre_tecnico FROM tecnologia_tecnico_responsable")
        proveedores_data = cur.fetchall()
        proveedores = {p["id"]: p["nombre_tecnico"] for p in proveedores_data}

        # Personas responsables → dict id:nombre
        cur.execute('SELECT id, nombre_contratista FROM tecnologia_persona_responsable')
        personas_data = cur.fetchall()
        personas = {r["id"]: r["nombre_contratista"] for r in personas_data}

        cur.execute('SELECT id, ubicacion_original FROM tecnologia_ubicacion_equipos')
        ubicacionEquipos_data = cur.fetchall()
        ubicacionEquipos = {p["id"]: p["ubicacion_original"] for p in ubicacionEquipos_data}

        historial = {
            'preventivo': preventivo
            # 'correctivo': correctivo
        }

        return render_template('historialPreventivoTecnologia.html', historial=historial, proveedores=proveedores, personas=personas, ubicacionEquipos=ubicacionEquipos)

    except Exception as e:
        print(f"Error al obtener el historial: {str(e)}")
        flash('Error al obtener el historial de fechas.', 'danger')
        return redirect(url_for('indexTecnologia'))
    finally:
        cur.close()


# HISTORIAL FECHAS MANTENIMIENTO CORRECTIVO TECNOLOGIA
@app.route('/historialCorrectivoTecnologia/<cod_articulo>')
@login_required
def HISTORIAL_CORRECTIVO_TECNOLOGIA(cod_articulo):
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        cur.execute(
            """
            SELECT id, cod_articulo, nombre_equipo, ubicacion_original, fecha_calibracion, 
                   vencimiento_calibracion, periodicidad_calibracion, id_proveedor_responsable, id_persona_responsable, observaciones
            FROM tecnologia_historial_correctivo 
            WHERE cod_articulo = %s 
            ORDER BY fecha_calibracion DESC
        """,
            [cod_articulo],
        )
        correctivo = cur.fetchall()

        # Trae el nombre de los técnicos
        cur.execute("SELECT id, nombre_tecnico FROM tecnologia_tecnico_responsable")
        proveedores_data = cur.fetchall()
        proveedores = {p["id"]: p["nombre_tecnico"] for p in proveedores_data}

        # Personas responsables → dict id:nombre
        cur.execute('SELECT id, nombre_contratista FROM tecnologia_persona_responsable')
        personas_data = cur.fetchall()
        personas = {r["id"]: r["nombre_contratista"] for r in personas_data}

        cur.execute('SELECT id, ubicacion_original FROM tecnologia_ubicacion_equipos')
        ubicacionEquipos_data = cur.fetchall()
        ubicacionEquipos = {p["id"]: p["ubicacion_original"] for p in ubicacionEquipos_data}

        historial = {
            # 'preventivo': preventivo,
            'correctivo': correctivo
        }

        return render_template('historialCorrectivoTecnologia.html', historial=historial, proveedores=proveedores, personas=personas, ubicacionEquipos=ubicacionEquipos)

    except Exception as e:
        print(f"Error al obtener el historial: {str(e)}")
        flash('Error al obtener el historial de fechas.', 'danger')
        return redirect(url_for('indexTecnologia'))
    finally:
        cur.close()


# ACTUALIZAR FECHAS DE MANTENIMIENTO PREVENTIVO
@app.route('/updateHistorialMantenimientoPreventivo', methods=['POST'])
def update_historial_mantenimiento_preventivo():
    id = request.form['id']  # ID del registro en historial_fechas
    cod_articulo = request.form['cod_articulo']
    fecha_mantenimiento = request.form['fecha_mantenimiento']
    vencimiento_mantenimiento = request.form['vencimiento_mantenimiento']
    periodicidad = request.form['periodicidad']

    # Usar DictCursor para obtener un diccionario
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Verificar si el registro es el último de su tipo en historial_fechas
    cur.execute("""SELECT id FROM tecnologia_historial_preventivo WHERE cod_articulo = %s ORDER BY id DESC LIMIT 1""", [cod_articulo])
    last_record = cur.fetchone()

   # Si el id del último registro coincide con el seleccionado, permitir la actualización
    if last_record and last_record['id'] == int(id):
        # Actualiza
        cur = db.connection.cursor()
        cur.execute("""UPDATE tecnologia_historial_preventivo SET fecha_mantenimiento = %s, 
                       vencimiento_mantenimiento = %s, periodicidad = %s WHERE id = %s""", 
                      (fecha_mantenimiento, vencimiento_mantenimiento, periodicidad, id))
        
        # Determinar el color del semáforo basado en la fecha de vencimiento más próxima
        fecha_actual = datetime.now().date()
        color = "verde"  # Valor por defecto

        # Comparar fechas de mantenimiento y calibración
        if vencimiento_mantenimiento:
            
            vencimiento_mant = datetime.strptime(vencimiento_mantenimiento, '%Y-%m-%d').date()
            if vencimiento_mant < fecha_actual + timedelta(days=0):
                color = "purple"
            elif vencimiento_mant <= fecha_actual + timedelta(days=30):
                color = "red"
            elif vencimiento_mant <= fecha_actual + timedelta(days=90):
                color = "yellow"
        
        db.connection.commit()
        flash('Historial actualizado correctamente', 'success')
    else:
        # Si no es el último registro, mostrar un mensaje de error
        flash('Solo se puede actualizar el último registro de este equipo.', 'danger')
    
    return redirect(url_for('historialPreventivoTecnologia', cod_articulo=cod_articulo))

# ACTUALIZAR FECHAS DE HISTORIAL MENTENIMIENTO CORRECTIVO
@app.route('/update_historialMantenimientoCorrectivo', methods=['POST'])
def update_historial_mantenimiento_correctivo():
    id = request.form['id']  # ID del registro en historial_fechas_calibracion
    cod_articulo = request.form['cod_articulo']
    fecha_calibracion = request.form['fecha_calibracion']
    vencimiento_calibracion = request.form['vencimiento_calibracion']
    periodicidad_calibracion = request.form['periodicidad_calibracion']

    # Usar DictCursor para obtener un diccionario
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Verificar si el registro es el último de su tipo en historial_fechas
    cur.execute("""SELECT id FROM tecnologia_historial_correctivo WHERE cod_articulo = %s ORDER BY id DESC LIMIT 1 """, [cod_articulo])
    last_record = cur.fetchone()

   # Si el id del último registro coincide con el seleccionado, permitir la actualización
    if last_record and last_record['id'] == int(id):
        # Actualizar
        cur = db.connection.cursor()
        cur.execute(""" UPDATE tecnologia_historial_correctivo SET fecha_calibracion = %s, 
                        vencimiento_calibracion = %s, periodicidad_calibracion = %s WHERE id = %s""", 
                        (fecha_calibracion, vencimiento_calibracion, periodicidad_calibracion, id))
        
        db.connection.commit()
        flash('Historial actualizado correctamente', 'success')
    else:
        # Si no es el último registro, mostrar un mensaje de error
        flash('Solo se puede actualizar el último registro de este equipo.', 'danger')
    
    return redirect(url_for('historialCorrectivoTecnologia', cod_articulo=cod_articulo))

# ======================================================================================================
# ==========================INICIA FUNCIÓN OTROS EQUIPOS DE TECNOLOGIA=====================
@app.route('/indexOtrosEquiposTecnologia')
@login_required
def index_otros_equipos_tecnologia():
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    # cur = db.connection.cursor()
    cur.execute('SELECT * FROM tecnologia_equipos WHERE enable=0 AND otros_equipos_tecnologia=1')
    data_otros_equipos_tecnologia = cur.fetchall()

    cur.execute('SELECT id, nombre_tecnico FROM tecnologia_tecnico_responsable')
    proveedores = cur.fetchall()

    cur.execute('SELECT id, documento_identidad, nombre_contratista FROM tecnologia_persona_responsable')
    personas = cur.fetchall()

    cur.execute('SELECT id, tipo_equipo FROM tecnologia_tipo_equipo')
    tipoEquipos = cur.fetchall()

    cur.execute('SELECT id, estado_equipo FROM tecnologia_estados_equipos')
    estadoEquipos = cur.fetchall()

    cur.execute('SELECT id, ubicacion_original FROM tecnologia_ubicacion_equipos')
    ubicacionEquipos = cur.fetchall()

    cur.execute('SELECT id, ubicacion_original FROM tecnologia_ubicacion_equipos')
    ubicacionEquipos_data = cur.fetchall()
    ubicacionEquiposModal = {p["id"]: p["ubicacion_original"] for p in ubicacionEquipos_data}
    # print(ubicacionEquipos)
    return render_template('indexOtrosEquiposTecnologia.html', tecnologia_equipos=data_otros_equipos_tecnologia, tipoEquipos=tipoEquipos, proveedores=proveedores, personas=personas, estadoEquipos=estadoEquipos, ubicacionEquipos=ubicacionEquipos, ubicacionEquiposModal=ubicacionEquiposModal)
# ==========================INICIA FUNCIÓN EQUIPOS DADOS DE BAJA TECNOLOGIA=====================
@app.route('/equiposDeBajaTecnologia')
@login_required
def equipos_debaja_tecnologia():
    
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute('SELECT id, estado_equipo FROM tecnologia_estados_equipos')
    estadoEquipos = cur.fetchall()
    
    cur.execute('SELECT * FROM tecnologia_equipos_debaja')
    equipos_de_baja_tecnologia = cur.fetchall()

    return render_template('equiposDeBajaTecnologia.html', equipos_de_baja_tecnologia=equipos_de_baja_tecnologia, estadoEquipos=estadoEquipos)
# ==========================FINALIZA FUNCIÓN EQUIPOS DADOS DE BAJA=====================

# FUNCIÓN ELIMINAR
@app.route('/delete_producto/<string:id>')
def ELIMINAR_CONTACTO(id):
    cur = db.connection.cursor()
    cur.execute('DELETE FROM productos WHERE id = {0}'.format(id))
    db.connection.commit()
    flash('Producto eliminado satisfactoriamente')
    return redirect(url_for('indexTecnologia'))
# --------------------------- FINALIZA MODULO DE TECNOLOGIA-----------------------------
# --------------------------- INICIA MODULO DE SALUD-----------------------------
# --------------------------- DATOS PROVEEDOR SALUD --------------------------------


# ---------------------------FUNCIÓN PARA EL MANEJO DE LOS MODULOS-----------------------------
@app.route('/<modulo>')
@login_required
def index_modulo(modulo):
    modulos_validos = ['salud', 'gastronomia', 'lacma', 'arquitectura']
    
    if modulo not in modulos_validos:
        # flash("Modulo no válido", "error")
        return redirect(url_for('home'))  # <-- redirige al home si no existe
    
    cur = db.connection.cursor(MySQLdb.cursors.DictCursor)

    # Traer equipos solo del modulo actual
    cur.execute("""SELECT i.*, p.enable_prestamos FROM indexssalud i LEFT JOIN prestamos_equiposalud p ON i.cod_articulo = p.cod_articulo AND p.enable_prestamos = 1 WHERE i.enable=1 AND i.de_baja=0 AND i.modulo=%s""", (modulo,))
    equipos = cur.fetchall()

    # Traer proveedores, estados y ubicaciones
    cur.execute('SELECT id, nombre_empresa FROM datosproveedorsalud')
    proveedores = cur.fetchall()

    cur.execute('SELECT id, estado_equipo FROM estados_equipos')
    estadoEquipos = cur.fetchall()

    cur.execute('SELECT id, ubicacion_original FROM ubicacion_equipos')
    ubicacionEquipos = cur.fetchall()

    return render_template(f'indexSalud.html', indexssalud=equipos, proveedores=proveedores, estadoEquipos=estadoEquipos, ubicacionEquipos=ubicacionEquipos, modulo=modulo)

    
# ---------------------------FUNCION PARA CARGAR IMAGEN DEL EQUIPO DESDE LA TABLA indexSalud EN EL CAMPO ACCIONES SUBIR_IMAGEN-----------------------------  
# @app.route('/subir_imagen/<int:id_producto>', methods=['POST'])
# def subir_imagen(id_producto, modulo):
#     if 'imagen_producto' not in request.files:
#         flash('No se seleccionó ningún archivo', 'error')
#         return redirect(url_for('index_modulo', modulo=modulo))

#     file = request.files['imagen_producto']
#     if file.filename == '':
#         flash('Por favor seleccione un archivo válido', 'error')
#         return redirect(url_for('index_modulo', modulo=modulo))
    
#     # Validar extensión
#     if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg')):
#         flash('Solo se permiten archivos PNG, JPG', 'error')
#         return redirect(url_for('index_modulo', modulo=modulo))

#     if file:
#         filename = secure_filename(file.filename)
#         filepath_to_db_img = os.path.join('fotos', filename).replace("\\", "/")
#         ruta_absoluta = os.path.join(app.root_path, 'static', filepath_to_db_img)

#         # Guardar en disco
#         file.save(ruta_absoluta)

#         # Actualizar en BD
#         cur = db.connection.cursor()
#         cur.execute("""
#             UPDATE indexssalud 
#             SET imagen = %s 
#             WHERE id = %s
#         """, (filepath_to_db_img, id_producto))
#         db.connection.commit()
#         cur.close()

#         flash('Imagen cargada correctamente', 'success')
#         return redirect(url_for('index_modulo', modulo='modulo'))

# ---------------------------FUNCION PARA CARGAR PDFS DEL EQUIPO DESDE LA TABLA indexSalud EN EL CAMPO ACCIONES SUBIR_GUIA---------------------------- 
# @app.route('/subir_pdf/<int:id_producto>', methods=['POST'])
# def subir_pdf(id_producto, modulo):
#     if 'pdf_salud' not in request.files:
#         flash('No se seleccionó ningún archivo', 'error')
#         return redirect(url_for('index_modulo', modulo=modulo))

#     file = request.files['pdf_salud']
#     if file.filename == '':
#         flash('Por favor seleccione un archivo válido', 'error')
#         return redirect(url_for('index_modulo', modulo=modulo))

#     # Validar extensión
#     if not file.filename.lower().endswith('.pdf'):
#         flash('Solo se permiten archivos PDF', 'error')
#         return redirect(url_for('index_modulo', modulo=modulo))

#     # Guardar archivo
#     filename = secure_filename(file.filename)
#     filepath_to_db_pdf = os.path.join('pdf', filename).replace("\\", "/")
#     ruta_absoluta = os.path.join(app.root_path, 'static', filepath_to_db_pdf)
#     file.save(ruta_absoluta)

#     # Actualizar en BD (columna ejemplo: pdf_salud)
#     cur = db.connection.cursor()
#     cur.execute("""
#         UPDATE indexssalud 
#         SET pdf_salud = %s 
#         WHERE id = %s
#     """, (filepath_to_db_pdf, id_producto))
#     db.connection.commit()
#     cur.close()

#     flash('Guia cargada correctamente', 'success')
#     return redirect(url_for('index_modulo', modulo='modulo'))

# ---------------------------INICIA INSERT MASIVO DE EQUIPOS CSV DE SALUD-----------------------------
# @app.route('/insert_csv', methods=['POST'])
# def insert_csv(modulo):
#     if request.method == 'POST':
#         file = request.files['file']
#         if not file:
#             flash('No seleccionó ningún archivo')
#             return redirect(url_for('index_modulo', modulo=modulo))

#         file = TextIOWrapper(file, encoding='latin-1')
#         csv_reader = csv.reader(file)
#         next(csv_reader)  # Saltar encabezado

#         cur = db.connection.cursor()
#         fotos_folder = os.path.join(os.path.dirname(__file__), 'static', 'fotos')

#         codigos_duplicados = []
#         datos_validos = []

#         for row in csv_reader:
#             cod_articulo = row[0]

#             try:
#                 cod_articulo = int(cod_articulo)
#             except ValueError:
#                 flash(f'Código inválido: {cod_articulo}', 'error')
#                 continue  # Salta esta fila

#             cur.execute("SELECT cod_articulo FROM indexssalud WHERE cod_articulo = %s", (cod_articulo,))
#             codigo_indexsalud = cur.fetchone()

#             cur.execute("SELECT cod_articulo FROM equipossalud_debaja WHERE cod_articulo = %s", (cod_articulo,))
#             codigo_debaja = cur.fetchone()

#             if codigo_indexsalud or codigo_debaja:
#                 codigos_duplicados.append(str(cod_articulo))
#                 continue  # Salta esta fila duplicada

#             # Verifica que la imagen exista
#             imagen = row[14]
#             imagen_path = os.path.join(fotos_folder, imagen)
#             if not os.path.isfile(imagen_path):
#                 flash(f'Imagen no encontrada: {imagen}', 'error')
#                 continue  # Salta esta fila

#             # Validación básica de Mantenimiento Actual
#             if not row[7]:
#                 flash(f'Equipo con código {cod_articulo} no tiene Mantenimiento Actual.', 'error')
#                 continue
            
#             # Validación básica de Vencimiento Mantenimiento
#             if not row[8]:
#                 flash(f'Equipo con código {cod_articulo} no tiene Vencimiento Mantenimiento.', 'error')
#                 continue

#             # Validación básica de Periodicidad Calibración
#             if not row[9]:
#                 flash(f'Equipo con código {cod_articulo} si no tiene periodicidad de calibracion, ingresa 0.', 'error')
#                 continue

#             # Preparar fila válida para insertar luego
#             datos_validos.append(row)

#             # Se construye diccionario de proveedores, por nombre de empresa: id
#             cur.execute("SELECT id, nombre_empresa FROM datosproveedorsalud")
#             proveedores = cur.fetchall()
#             proveedor_map = {p[1].strip().lower(): p[0] for p in proveedores}

#         # Insertar solo los datos válidos
#         for row in datos_validos:
#             cod_articulo = int(row[0])
#             nombre_equipo = row[1]
#             ubicacion_original = row[2]
#             estado_equipo = row[3]
#             fecha_ingreso = row[4]
#             garantia = row[5]
#             periodicidad = int(row[6])
#             fecha_mantenimiento = row[7]
#             vencimiento_mantenimiento = row[8]
#             periodicidad_calibracion = int(row[9])
#             fecha_calibracion = row[10] or None
#             vencimiento_calibracion = row[11] or None
#             criticos = row[12]
#             # proveedor_responsable = row[13]
#             nombre_proveedor = row[13].strip().lower()
#             proveedor_responsable = proveedor_map.get(nombre_proveedor)

#             if not proveedor_responsable:
#                 flash(f"Proveedor '{row[13]}' no encontrado en la base de datos.", 'error')
#                 continue
            
#             imagen = row[14]
#             especificaciones_instalacion = row[15]
#             cuidados_basicos = row[16]
#             marca_equipo_salud = row[17]
#             modelo_equipo_salud = row[18]
#             serial_equipo_salud = row[19]

#             ruta_imagen_db = f'fotos/{secure_filename(imagen)}'
#             checkbox_mantenimiento = 'Inactivo'
#             checkbox_calibracion = 'Inactivo'
#             fecha_de_baja = date.today() if estado_equipo == "DE BAJA" else None
#             color = 'verde'

#             if estado_equipo == 'DE BAJA':
#                 # Insertar en la tabla equipossalud_debaja
#                 cur.execute("""INSERT INTO equipossalud_debaja (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, fecha_ingreso, 
#                                                                 periodicidad, estado_equipo, ubicacion_original, garantia, criticos, proveedor_responsable, imagen, especificaciones_instalacion, cuidados_basicos, 
#                                                                 periodicidad_calibracion, marca_equipo_salud, modelo_equipo_salud, serial_equipo_salud, color, checkbox_mantenimiento, checkbox_calibracion, fecha_de_baja) VALUES 
#                                                                 (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
#                     (
#                         cod_articulo,
#                         nombre_equipo,
#                         fecha_mantenimiento,
#                         vencimiento_mantenimiento,
#                         fecha_calibracion,
#                         vencimiento_calibracion,
#                         fecha_ingreso,
#                         periodicidad,
#                         estado_equipo,
#                         ubicacion_original,
#                         garantia,
#                         criticos,
#                         proveedor_responsable,
#                         ruta_imagen_db,
#                         especificaciones_instalacion, 
#                         cuidados_basicos,
#                         periodicidad_calibracion,
#                         marca_equipo_salud,
#                         modelo_equipo_salud,
#                         serial_equipo_salud,
#                         color,
#                         checkbox_mantenimiento,
#                         checkbox_calibracion,
#                         fecha_de_baja,
#                     ),
#                 )
#             else:

#                 # Insertar en la tabla indexssalud
#                 cur.execute("""INSERT INTO indexssalud (cod_articulo, nombre_equipo, fecha_mantenimiento, vencimiento_mantenimiento, fecha_calibracion, vencimiento_calibracion, fecha_ingreso, 
#                                                         periodicidad, estado_equipo, ubicacion_original, garantia, criticos, proveedor_responsable, imagen, especificaciones_instalacion, cuidados_basicos, 
#                                                         periodicidad_calibracion, marca_equipo_salud, modelo_equipo_salud, serial_equipo_salud, color, checkbox_mantenimiento, checkbox_calibracion, fecha_de_baja) VALUES 
#                                                         (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
#                     (
#                         cod_articulo,
#                         nombre_equipo,
#                         fecha_mantenimiento,
#                         vencimiento_mantenimiento,
#                         fecha_calibracion,  
#                         vencimiento_calibracion,  
#                         fecha_ingreso,
#                         periodicidad,
#                         estado_equipo,
#                         ubicacion_original,
#                         garantia,
#                         criticos,
#                         proveedor_responsable,
#                         ruta_imagen_db,
#                         especificaciones_instalacion, 
#                         cuidados_basicos,
#                         periodicidad_calibracion,
#                         marca_equipo_salud,
#                         modelo_equipo_salud,
#                         serial_equipo_salud,
#                         color,
#                         checkbox_mantenimiento,
#                         checkbox_calibracion,
#                         fecha_de_baja,
#                     ),
#                 )

#         db.connection.commit()

#         if codigos_duplicados:
#             flash(f'Los siguientes códigos ya existen y no fueron insertados: {", ".join(codigos_duplicados)}', 'error')
#         if datos_validos:
#             flash(f'{len(datos_validos)} equipos importados exitosamente.', 'success')
#         else:
#             flash('No se importó ningún equipo.', 'error')

#         return redirect(url_for('index_modulo', modulo='modulo'))
# ---------------------------FINALIZA INSERT MASIVO CSV SALUD-----------------------------

# ---------------------------INICIA EXPORTACIÓN DE CSV DE EQUIPOS TECNOLOGIA-----------------------------
@app.route('/exportCsvTecnologia')
@login_required
def exportCsv():
    cur = db.connection.cursor()

    # Consulta SQL optimizada con JOINs correctos y alias claros
    cur.execute("""
        SELECT 
            i.cod_articulo,
            i.nombre_equipo,
            i.fecha_ingreso,
            i.fecha_mantenimiento,
            i.vencimiento_mantenimiento,
            i.fecha_calibracion,
            i.vencimiento_calibracion,
            i.estado_equipo,
            u.ubicacion_original AS proceso,
            i.marca_equipo_tecnologia,
            i.modelo_equipo_tecnologia,
            i.serial_equipo_tecnologia,
            i.ram,
            i.disco,
            i.tipo_equipo,
            i.software_instalado,
            t.nombre_tecnico AS proveedor_responsable,
            q.documento_identidad AS id_persona_responsable,
            p.nombre_contratista AS id_persona_responsable
            # i.enable
        FROM tecnologia_equipos i
        LEFT JOIN tecnologia_tecnico_responsable t ON i.proveedor_responsable = t.id
        LEFT JOIN tecnologia_persona_responsable q ON i.id_persona_responsable = q.id
        LEFT JOIN tecnologia_persona_responsable p ON i.id_persona_responsable = p.id
        LEFT JOIN tecnologia_ubicacion_equipos u ON i.ubicacion_original = u.id
        WHERE i.de_baja = 0
        ORDER BY i.cod_articulo ASC
    """)

    registros = cur.fetchall()
    cur.close()

    # Crear el archivo CSV en memoria
    si = StringIO()
    writer = csv.writer(si)

    # Encabezados claros y ordenados
    encabezados = [
        'Código Equipo',
        'Nombre Equipo',
        'Fecha Ingreso',
        'Fecha Ejecución Preventivo',
        'Fecha Vencimiento Preventivo',
        'Fecha Ejecución Correctivo',
        'Fecha Vencimiento Correctivo',
        'Estado Equipo',
        'Proceso',
        'Marca Equipo',
        'Modelo Equipo',
        'Serial Equipo',
        'Memoria RAM',
        'Disco Duro',
        'Tipo Equipo',
        'Software Instalado',
        'Técnico Responsable',
        'ID Responsable del Equipo',
        'Nombre Responsable del Equipo',
        # 'Estado Equipo'
    ]
    writer.writerow(encabezados)

    # Escribir los registros de la tabla
    for registro in registros:
        writer.writerow(registro)

    # Preparar la respuesta de descarga
    salida = Response(
        si.getvalue().encode('utf-8-sig'),
        mimetype='text/csv'
    )
    salida.headers['Content-Disposition'] = 'attachment; filename=equiposTecnologia.csv'

    return salida
# ---------------------------FINALIZA EXPORTACIÓN DE CSV DE EQUIPOS DE SALUD-----------------------------

# ---------------------------INICIA EXPORTACIÓN DE FORMATO EXCEL DE EQUIPOS DE BAJA DE TECNOLIGIA-----------------------------
@app.route('/exportExcelTecnologiaDeBaja', methods=['POST'])
@login_required
def exportExcelDeBaja():
    try:
        # 1) Validar JSON/csrf
        if not request.is_json:
            return jsonify({"error": "El cuerpo de la petición debe ser JSON"}), 400

        data = request.get_json(silent=True) or {}
        equipos = data.get('equipos') or []

        if not isinstance(equipos, list) or not equipos:
            return jsonify({"error": "No se enviaron equipos"}), 400

        # 2) Ubicar plantilla
        plantilla_path = os.path.join(current_app.root_path, "static", "img", "INFORME_TECNICO_BAJAS.xlsx")
        if not os.path.exists(plantilla_path):
            return jsonify({"error": f"No se encontró la plantilla en {plantilla_path}"}), 400

        # 3) Cargar plantilla
        wb = load_workbook(plantilla_path)
        ws = wb.active  # Ajusta a wb['NombreHoja'] si tu hoja no es la activa

        # 4) Escribir datos desde fila 12 (debajo de los títulos)
        start_row = 12

        # Helper: verificar si un rango ya está fusionado
        def is_merged(ws, coord: str) -> bool:
            return any(str(r) == coord for r in ws.merged_cells.ranges)

        for idx, equipo in enumerate(equipos, start=start_row):
            placa = (equipo.get("cod_articulo") or "").strip()
            cantidad = 1
            nombre = (equipo.get("nombre_equipo") or "").strip()

            # Columna A → PLACA
            ws[f"A{idx}"].value = placa
                       
            # Columna B → CANTIDAD
            ws[f"B{idx}"].value = cantidad

            # Columnas C-G → DESCRIPCIÓN DEL BIEN (escribir en C, y fusionar sólo si hace falta)
            merge_coord = f"C{idx}:G{idx}"
            if not is_merged(ws, merge_coord):
                try:
                    ws.merge_cells(merge_coord)
                except ValueError:
                    # Si ya está fusionado por plantilla o solapa, lo ignoramos
                    pass

            ccell = ws[f"C{idx}"]
            ccell.value = nombre
            ccell.alignment = Alignment(wrap_text=True, vertical="top")

        # 5) Enviar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="equipos_baja.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        current_app.logger.exception("Error generando Excel")
        return jsonify({"error": str(e)}), 400
# ---------------------------FINALIZA EXPORTACIÓN DE CSV DE EQUIPOS DE EQUIPOS DE BAJA DE TECNOLIGIA-----------------------------

# ==========================FINALIZA FUNCIÓN EQUIPOS DADOS DE BAJA SALUD=====================


# FUNCIÓN ELIMINAR PARA INDEXSSALUD
# @app.route('/delete_productoSalud/<string:id>')
# @login_required
# def ELIMINAR_CONTACTO_SALUD(id):
#     cur = db.connection.cursor()
#     # cur.execute('DELETE FROM indexssalud WHERE id = {0}'.format(id))
#     #Esta linea de codigo en la vista elimina el producto pero no de la DB, la cual realiza es una actualización
#     cur.execute('UPDATE indexssalud SET enable=0 WHERE id = {0}'.format(id))
#     db.connection.commit()
#     flash('Equipo eliminado satisfactoriamente', 'success')
#     return redirect(url_for('index_modulo', modulo='modulo'))


def status_401(error):
    return redirect(url_for('login'))


def status_404(error):
    return "<h1>Página no encontrada</h1>", 404

app.register_error_handler(401, status_401)
app.register_error_handler(404, status_404)


if __name__ == '__main__':

    app.run()
